"""
P6 XER Project Manager Planning Tool
=====================================
A Streamlit app for interrogating Primavera P6 XER schedules
without needing to open P6. Designed for Project Managers.
"""

import io
import re
import math
import warnings
from collections import defaultdict, deque
from datetime import datetime, timedelta

import networkx as nx
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# -----------------------------------------------------------------------------
# PAGE CONFIG
# -----------------------------------------------------------------------------
st.set_page_config(
    page_title="PlanTrace",
    page_icon="🎯",
    layout="wide",
    initial_sidebar_state="expanded",
)

# -----------------------------------------------------------------------------
# CUSTOM CSS  --  PlanTrace brand palette
# -----------------------------------------------------------------------------
st.markdown("""
<style>
    /* ---- Global background ---- */
    .stApp { background-color: #F5F7FA; }

    /* ---- Sidebar ---- */
    [data-testid="stSidebar"] {
        background-color: #0B1F33;
    }
    [data-testid="stSidebar"] * { color: #CBD5E1 !important; }
    [data-testid="stSidebar"] .sidebar-brand {
        color: #F5A623 !important; font-weight: 800;
    }
    [data-testid="stSidebar"] hr { border-color: #1e3a5f; }

    /* ---- Headings ---- */
    h1, h2, h3 { color: #0B1F33; font-weight: 700; }
    h4, h5     { color: #1e3a5f; }

    /* ---- Metric cards ---- */
    div[data-testid="metric-container"] {
        background: #ffffff;
        border-radius: 10px;
        padding: 14px 18px;
        border: 1px solid #E2E8F0;
        box-shadow: 0 1px 4px rgba(11,31,51,0.07);
    }

    /* ---- DataFrames ---- */
    .stDataFrame { border-radius: 8px; }

    /* ---- Utility boxes ---- */
    .pt-warn {
        background: #fffbeb; border-left: 4px solid #F5A623;
        border-radius: 6px; padding: 12px 16px; margin: 8px 0;
    }
    .pt-info {
        background: #eff6ff; border-left: 4px solid #0B1F33;
        border-radius: 6px; padding: 12px 16px; margin: 8px 0;
    }
    .pt-error {
        background: #fef2f2; border-left: 4px solid #dc2626;
        border-radius: 6px; padding: 12px 16px; margin: 8px 0;
    }
    .pt-ok {
        background: #f0fdf4; border-left: 4px solid #16a34a;
        border-radius: 6px; padding: 12px 16px; margin: 8px 0;
    }

    /* ---- Feature cards on homepage ---- */
    .pt-card {
        background: #ffffff;
        border-radius: 12px;
        padding: 28px 24px;
        border: 1px solid #E2E8F0;
        box-shadow: 0 2px 8px rgba(11,31,51,0.08);
        height: 100%;
        transition: box-shadow 0.2s;
    }
    .pt-card:hover {
        box-shadow: 0 4px 16px rgba(11,31,51,0.14);
    }
    .pt-card-icon {
        font-size: 32px; margin-bottom: 12px;
    }
    .pt-card-title {
        font-size: 17px; font-weight: 700;
        color: #0B1F33; margin-bottom: 8px;
    }
    .pt-card-body {
        font-size: 14px; color: #64748B; line-height: 1.6;
    }
    .pt-card-accent {
        display: inline-block;
        width: 36px; height: 3px;
        background: #F5A623;
        border-radius: 2px;
        margin-bottom: 14px;
    }

    /* ---- Badges ---- */
    .badge-critical  { background:#dc2626; color:#fff; padding:2px 9px;
                       border-radius:12px; font-size:11px; font-weight:700; }
    .badge-amber     { background:#F5A623; color:#fff; padding:2px 9px;
                       border-radius:12px; font-size:11px; font-weight:700; }
    .badge-ok        { background:#16a34a; color:#fff; padding:2px 9px;
                       border-radius:12px; font-size:11px; font-weight:700; }
    .badge-navy      { background:#0B1F33; color:#fff; padding:2px 9px;
                       border-radius:12px; font-size:11px; font-weight:700; }

    /* ---- Upload prompt ---- */
    .pt-upload-box {
        background: #ffffff;
        border: 2px dashed #CBD5E1;
        border-radius: 12px;
        padding: 32px 24px;
        text-align: center;
        color: #64748B;
    }
</style>
""", unsafe_allow_html=True)


# -----------------------------------------------------------------------------
# XER PARSING  (xerparser + manual fallback)
# -----------------------------------------------------------------------------

def parse_xer_fallback(raw_text: str) -> dict:
    """
    Manual fallback parser that reads XER table format:
    %T TABLE_NAME  /  %F col1 col2 ...  /  %R val1 val2 ...
    Returns dict of {table_name: list_of_dicts}
    """
    tables = {}
    current_table = None
    current_fields = []

    for line in raw_text.splitlines():
        line = line.rstrip("\r")
        if line.startswith("%T\t"):
            current_table = line[3:].strip()
            current_fields = []
            tables[current_table] = []
        elif line.startswith("%F\t") and current_table:
            current_fields = line[3:].split("\t")
        elif line.startswith("%R\t") and current_table and current_fields:
            values = line[3:].split("\t")
            # Pad values if shorter than fields
            while len(values) < len(current_fields):
                values.append("")
            row = {current_fields[i]: values[i] for i in range(len(current_fields))}
            tables[current_table].append(row)

    return tables


def hours_to_days(hours, hours_per_day=8.0):
    """Convert hours to working days."""
    if hours is None:
        return None
    try:
        return round(float(hours) / hours_per_day, 1)
    except (TypeError, ValueError):
        return None


def safe_float(val, default=None):
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


def safe_date(val):
    if val is None or str(val).strip() in ("", "None"):
        return None
    if isinstance(val, datetime):
        return val
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d", "%d/%m/%Y %H:%M", "%d/%m/%Y"):
        try:
            return datetime.strptime(str(val).strip(), fmt)
        except ValueError:
            pass
    return None


def parse_xer(file_bytes: bytes):
    """
    Parse an XER file. Uses xerparser library first; falls back to manual parsing.
    Returns a dict with keys: tasks_df, relationships_df, wbs_df, resources_df,
    task_resources_df, project_info, calendars_df, parse_method
    """
    # Try to decode the file
    for codec in ("cp1252", "utf-8", "latin-1"):
        try:
            raw_text = file_bytes.decode(codec)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("Cannot decode XER file. Please check the file encoding.")

    result = {
        "tasks_df": pd.DataFrame(),
        "relationships_df": pd.DataFrame(),
        "wbs_df": pd.DataFrame(),
        "resources_df": pd.DataFrame(),
        "task_resources_df": pd.DataFrame(),
        "project_info": {},
        "calendars_df": pd.DataFrame(),
        "parse_method": "unknown",
    }

    # -- Try xerparser library -------------------------------------------------
    try:
        from xerparser.src.xer import Xer
        xer = Xer(raw_text)

        # Project info
        proj = None
        if xer.projects:
            proj_id = next(iter(xer.projects))
            proj = xer.projects[proj_id]
            result["project_info"] = {
                "name": getattr(proj, "name", ""),
                "data_date": getattr(proj, "last_recalc_date", None),
                "project_id": proj_id,
                "plan_start": getattr(proj, "plan_start_date", None),
                "scd_end": getattr(proj, "scd_end_date", None),
            }

        # Tasks DataFrame
        rows = []
        for uid, task in xer.tasks.items():
            tf = task.total_float_hr_cnt
            ff = task.free_float_hr_cnt
            # Effective start/finish (actual if done, early if not)
            eff_start = task.act_start_date or task.early_start_date or task.target_start_date
            eff_finish = task.act_end_date or task.early_end_date or task.target_end_date

            # WBS path
            wbs_node = xer.wbs_nodes.get(task.wbs_id)
            wbs_path = ""
            if wbs_node:
                parts = []
                n = wbs_node
                while n:
                    parts.append(getattr(n, "name", ""))
                    n = getattr(n, "parent", None)
                wbs_path = " > ".join(reversed(parts))

            # Calendar name
            cal = xer.calendars.get(task.clndr_id)
            cal_name = getattr(cal, "name", "") if cal else ""

            rows.append({
                "task_id": uid,
                "task_code": task.task_code,
                "task_name": task.name,
                "wbs_id": task.wbs_id,
                "wbs_path": wbs_path,
                "status": task.status.value if task.status else "",
                "task_type": task.type.value if task.type else "",
                "calendar": cal_name,
                "early_start": task.early_start_date,
                "early_finish": task.early_end_date,
                "late_start": task.late_start_date,
                "late_finish": task.late_end_date,
                "act_start": task.act_start_date,
                "act_finish": task.act_end_date,
                "target_start": task.target_start_date,
                "target_finish": task.target_end_date,
                "eff_start": eff_start,
                "eff_finish": eff_finish,
                "orig_dur_days": hours_to_days(task.target_drtn_hr_cnt),
                "rem_dur_days": hours_to_days(task.remain_drtn_hr_cnt),
                "total_float_days": hours_to_days(tf),
                "free_float_days": hours_to_days(ff),
                "total_float_hrs": tf,
                "is_longest_path": task.is_longest_path,
                "cstr_type": task.cstr_type,
                "cstr_date": task.cstr_date,
                "cstr_type2": task.cstr_type2,
                "cstr_date2": task.cstr_date2,
                "phys_pct": round(task.phys_complete_pct * 100, 1),
                "float_path": task.float_path,
            })

        result["tasks_df"] = pd.DataFrame(rows)

        # Relationships DataFrame
        rel_rows = []
        for uid, rel in xer.relationships.items():
            rel_rows.append({
                "pred_id": uid,
                "pred_task_id": rel.predecessor.uid if rel.predecessor else "",
                "pred_task_code": rel.predecessor.task_code if rel.predecessor else "",
                "pred_task_name": rel.predecessor.name if rel.predecessor else "",
                "succ_task_id": rel.successor.uid if rel.successor else "",
                "succ_task_code": rel.successor.task_code if rel.successor else "",
                "succ_task_name": rel.successor.name if rel.successor else "",
                "rel_type": rel.link,
                "lag_days": rel.lag,
                "lag_hrs": rel.lag_hr_cnt,
            })
        result["relationships_df"] = pd.DataFrame(rel_rows)

        # WBS DataFrame
        wbs_rows = []
        for uid, wbs in xer.wbs_nodes.items():
            wbs_rows.append({
                "wbs_id": uid,
                "wbs_code": getattr(wbs, "short_name", ""),
                "wbs_name": getattr(wbs, "name", ""),
                "parent_wbs_id": getattr(wbs, "parent_wbs_id", ""),
                "proj_id": getattr(wbs, "proj_id", ""),
            })
        result["wbs_df"] = pd.DataFrame(wbs_rows)

        # Resources & task resources
        if xer.resources:
            res_rows = []
            for uid, r in xer.resources.items():
                res_rows.append({
                    "rsrc_id": uid,
                    "rsrc_name": getattr(r, "name", ""),
                    "rsrc_short": getattr(r, "rsrc_short_name", ""),
                    "rsrc_type": getattr(r, "rsrc_type", ""),
                })
            result["resources_df"] = pd.DataFrame(res_rows)

        # Task resources (loading)
        taskrsrc_rows = []
        for uid, task in xer.tasks.items():
            for tr in getattr(task, "resources", []):
                taskrsrc_rows.append({
                    "task_id": uid,
                    "task_code": task.task_code,
                    "rsrc_id": getattr(tr, "rsrc_id", ""),
                    "target_qty": safe_float(getattr(tr, "target_qty", 0), 0),
                    "remain_qty": safe_float(getattr(tr, "remain_qty", 0), 0),
                    "act_reg_qty": safe_float(getattr(tr, "act_reg_qty", 0), 0),
                    "target_start": safe_date(getattr(tr, "target_start_date", None)),
                    "target_finish": safe_date(getattr(tr, "target_end_date", None)),
                })
        result["task_resources_df"] = pd.DataFrame(taskrsrc_rows)

        result["parse_method"] = "xerparser"
        return result

    except Exception as e:
        st.warning(f"xerparser failed ({e}), using fallback parser...")

    # -- Manual fallback -------------------------------------------------------
    try:
        tables = parse_xer_fallback(raw_text)
        return _build_from_raw_tables(tables)
    except Exception as e2:
        raise ValueError(f"Both parsers failed. Last error: {e2}")


def _build_from_raw_tables(tables: dict) -> dict:
    """Build result dict from raw parsed tables (fallback)."""
    result = {
        "tasks_df": pd.DataFrame(),
        "relationships_df": pd.DataFrame(),
        "wbs_df": pd.DataFrame(),
        "resources_df": pd.DataFrame(),
        "task_resources_df": pd.DataFrame(),
        "project_info": {},
        "calendars_df": pd.DataFrame(),
        "parse_method": "manual_fallback",
    }

    # Project info
    if "PROJECT" in tables and tables["PROJECT"]:
        proj = tables["PROJECT"][0]
        result["project_info"] = {
            "name": proj.get("proj_short_name", proj.get("proj_id", "")),
            "data_date": safe_date(proj.get("last_recalc_date")),
            "plan_start": safe_date(proj.get("plan_start_date")),
            "scd_end": safe_date(proj.get("scd_end_date")),
        }

    # Tasks
    if "TASK" in tables:
        df = pd.DataFrame(tables["TASK"])
        # Normalise date columns
        for col in ["early_start_date", "early_end_date", "late_start_date",
                    "late_end_date", "act_start_date", "act_end_date",
                    "target_start_date", "target_end_date", "cstr_date", "cstr_date2"]:
            if col in df.columns:
                df[col] = df[col].apply(safe_date)
        # Float
        for col in ["total_float_hr_cnt", "free_float_hr_cnt",
                    "target_drtn_hr_cnt", "remain_drtn_hr_cnt"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")

        # Build normalised columns
        df["eff_start"] = df.get("act_start_date", df.get("early_start_date"))
        df["eff_finish"] = df.get("act_end_date", df.get("early_end_date"))
        df["total_float_days"] = df.get("total_float_hr_cnt", pd.Series(dtype=float)).apply(hours_to_days)
        df["free_float_days"] = df.get("free_float_hr_cnt", pd.Series(dtype=float)).apply(hours_to_days)
        df["orig_dur_days"] = df.get("target_drtn_hr_cnt", pd.Series(dtype=float)).apply(hours_to_days)
        df["rem_dur_days"] = df.get("remain_drtn_hr_cnt", pd.Series(dtype=float)).apply(hours_to_days)

        # Rename for consistency
        rename = {
            "task_id": "task_id", "task_code": "task_code",
            "task_name": "task_name", "wbs_id": "wbs_id",
            "status_code": "status", "task_type": "task_type",
            "early_start_date": "early_start", "early_end_date": "early_finish",
            "late_start_date": "late_start", "late_end_date": "late_finish",
            "act_start_date": "act_start", "act_end_date": "act_finish",
            "target_start_date": "target_start", "target_end_date": "target_finish",
            "cstr_type": "cstr_type", "cstr_date": "cstr_date",
            "cstr_type2": "cstr_type2", "cstr_date2": "cstr_date2",
            "driving_path_flag": "is_longest_path_flag",
            "phys_complete_pct": "phys_pct",
        }
        df = df.rename(columns={k: v for k, v in rename.items() if k in df.columns})
        if "is_longest_path_flag" in df.columns:
            df["is_longest_path"] = df["is_longest_path_flag"] == "Y"
        df["wbs_path"] = df.get("wbs_id", "")
        result["tasks_df"] = df

    # Relationships
    if "TASKPRED" in tables:
        df = pd.DataFrame(tables["TASKPRED"])
        if "lag_hr_cnt" in df.columns:
            df["lag_days"] = pd.to_numeric(df["lag_hr_cnt"], errors="coerce").apply(hours_to_days)
        rename_r = {"pred_type": "rel_type", "task_id": "succ_task_id",
                    "pred_task_id": "pred_task_id"}
        df = df.rename(columns={k: v for k, v in rename_r.items() if k in df.columns})
        result["relationships_df"] = df

    # WBS
    if "PROJWBS" in tables:
        df = pd.DataFrame(tables["PROJWBS"])
        rename_w = {"wbs_id": "wbs_id", "wbs_short_name": "wbs_code",
                    "wbs_name": "wbs_name", "parent_wbs_id": "parent_wbs_id"}
        df = df.rename(columns={k: v for k, v in rename_w.items() if k in df.columns})
        result["wbs_df"] = df

    # Resources
    if "RSRC" in tables:
        df = pd.DataFrame(tables["RSRC"])
        rename_rs = {"rsrc_id": "rsrc_id", "rsrc_name": "rsrc_name",
                     "rsrc_short_name": "rsrc_short"}
        df = df.rename(columns={k: v for k, v in rename_rs.items() if k in df.columns})
        result["resources_df"] = df

    # Task resources
    if "TASKRSRC" in tables:
        df = pd.DataFrame(tables["TASKRSRC"])
        for col in ["target_qty", "remain_qty", "act_reg_qty"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
        for col in ["target_start_date", "target_end_date"]:
            if col in df.columns:
                df[col] = df[col].apply(safe_date)
                df = df.rename(columns={col: col.replace("_date", "")})
        result["task_resources_df"] = df

    return result


# -----------------------------------------------------------------------------
# GRAPH BUILDING
# -----------------------------------------------------------------------------

def build_graph(tasks_df: pd.DataFrame, rels_df: pd.DataFrame) -> nx.DiGraph:
    """Build a networkx directed graph from tasks and relationships."""
    G = nx.DiGraph()
    for _, row in tasks_df.iterrows():
        G.add_node(row["task_id"], **row.to_dict())
    for _, row in rels_df.iterrows():
        if row.get("pred_task_id") and row.get("succ_task_id"):
            G.add_edge(
                row["pred_task_id"],
                row["succ_task_id"],
                rel_type=row.get("rel_type", "FS"),
                lag_days=row.get("lag_days", 0),
            )
    return G


# -----------------------------------------------------------------------------
# CRITICAL PATH HELPERS
# -----------------------------------------------------------------------------

def get_critical_threshold(tasks_df: pd.DataFrame, near_crit_days: float = 10.0):
    """Classify activities as critical / near-critical / float."""
    df = tasks_df.copy()
    df["is_critical"] = df["total_float_days"].apply(
        lambda f: f is not None and f <= 0
    )
    df["is_near_critical"] = df["total_float_days"].apply(
        lambda f: f is not None and 0 < f <= near_crit_days
    )
    return df


def float_status_badge(f):
    if f is None:
        return "-"
    elif f <= 0:
        return "🔴 Critical"
    elif f <= 10:
        return "🟡 Near-Critical"
    else:
        return "🟢 Float"


# -----------------------------------------------------------------------------
# LOGIC TRACE HELPERS
# -----------------------------------------------------------------------------

def trace_predecessors(G: nx.DiGraph, task_id: str, max_depth=100) -> list:
    """BFS backwards through predecessors. Returns list of (task_id, depth)."""
    visited = {}
    queue = deque([(task_id, 0)])
    result = []
    while queue:
        node, depth = queue.popleft()
        if node in visited or depth > max_depth:
            continue
        visited[node] = depth
        if node != task_id:
            result.append((node, depth))
        for pred in G.predecessors(node):
            if pred not in visited:
                queue.append((pred, depth + 1))
    return result


def trace_successors(G: nx.DiGraph, task_id: str, max_depth=100) -> list:
    """BFS forwards through successors."""
    visited = {}
    queue = deque([(task_id, 0)])
    result = []
    while queue:
        node, depth = queue.popleft()
        if node in visited or depth > max_depth:
            continue
        visited[node] = depth
        if node != task_id:
            result.append((node, depth))
        for succ in G.successors(node):
            if succ not in visited:
                queue.append((succ, depth + 1))
    return result


def driving_path_to_activity(
    G: nx.DiGraph,
    tasks_df: pd.DataFrame,
    rels_df: pd.DataFrame,
    target_id: str,
) -> list:
    """
    Identify the most likely driving predecessor chain into a target activity.

    Driving predecessor selection priority (in order):
      1. Lowest total float  (most constrained activity wins)
      2. On P6 longest-path / driving flag where available
      3. Latest early-finish date  (latest predecessor is usually the driver)
      4. Highest lag on the connecting relationship (more constraining)

    Returns ordered list of task_ids, from chain start -> target.
    """
    task_lookup = tasks_df.set_index("task_id").to_dict("index") if not tasks_df.empty else {}

    def _score(pred_id, succ_id):
        t  = task_lookup.get(pred_id, {})
        tf = safe_float(t.get("total_float_days"), 9999)
        finish = t.get("eff_finish")
        if finish is not None:
            try:
                finish_score = -(finish.timestamp() / 86400)
            except Exception:
                finish_score = 0
        else:
            finish_score = 0
        driving_bonus = 0 if t.get("is_longest_path", False) else 1
        lag_score = 0
        if not rels_df.empty:
            rel = rels_df[
                (rels_df.get("pred_task_id", pd.Series(dtype=str)) == pred_id) &
                (rels_df.get("succ_task_id", pd.Series(dtype=str)) == succ_id)
            ]
            if not rel.empty and "lag_days" in rel.columns:
                lag_val = safe_float(rel["lag_days"].iloc[0], 0)
                lag_score = -lag_val
        return (tf, driving_bonus, finish_score, lag_score)

    path    = [target_id]
    visited = {target_id}
    current = target_id

    for _ in range(500):
        preds = list(G.predecessors(current))
        if not preds:
            break
        unvisited = [p for p in preds if p not in visited]
        if not unvisited:
            break
        best = min(unvisited, key=lambda p: _score(p, current))
        path.insert(0, best)
        visited.add(best)
        current = best

    return path


def _all_pred_paths(
    G: nx.DiGraph,
    tasks_df: pd.DataFrame,
    target_id: str,
    max_paths: int = 8,
) -> list:
    """
    Find up to max_paths predecessor chains into target_id.
    Each chain is a list of task_ids ordered start -> target.
    Only returns chains that begin at an activity with no predecessors.
    """
    task_lookup = tasks_df.set_index("task_id").to_dict("index") if not tasks_df.empty else {}

    def _float(tid):
        return safe_float(task_lookup.get(tid, {}).get("total_float_days"), 9999)

    found_paths = []

    def dfs(node, current_path, visited_set):
        if len(found_paths) >= max_paths:
            return
        preds = [p for p in G.predecessors(node) if p not in visited_set]
        if not preds:
            found_paths.append(list(reversed(current_path)))
            return
        for pred in sorted(preds, key=_float)[:4]:
            dfs(pred, current_path + [pred], visited_set | {pred})

    dfs(target_id, [target_id], {target_id})
    return found_paths


# -----------------------------------------------------------------------------
# EXPORT HELPERS
# -----------------------------------------------------------------------------

def style_header_row(ws, row_idx, fill_color="1e3a5f", font_color="FFFFFF"):
    fill = PatternFill("solid", start_color=fill_color, fgColor=fill_color)
    font = Font(bold=True, color=font_color)
    for cell in ws[row_idx]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def df_to_sheet(ws, df, sheet_title=None):
    """Write a DataFrame to an openpyxl worksheet with formatting."""
    if sheet_title:
        ws.title = sheet_title[:31]
    ws.append(list(df.columns))
    style_header_row(ws, 1)
    for r in df.itertuples(index=False):
        ws.append(list(r))
    # Auto-width
    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 4, 50)


def export_df_to_excel(sheets: dict) -> bytes:
    """sheets = {sheet_name: dataframe}. Returns Excel bytes."""
    wb = Workbook()
    first = True
    for name, df in sheets.items():
        if first:
            ws = wb.active
            first = False
        else:
            ws = wb.create_sheet()
        df_to_sheet(ws, df, name)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def format_date(d):
    if d is None:
        return "-"
    try:
        return d.strftime("%d %b %Y")
    except Exception:
        return str(d)


# -----------------------------------------------------------------------------
# PAGE: PROJECT SUMMARY
# -----------------------------------------------------------------------------

def page_project_summary(data: dict, near_crit_days: float):
    st.title("📊 Project Summary")

    proj = data["project_info"]
    tasks = data["tasks_df"]
    rels = data["relationships_df"]

    if tasks.empty:
        st.warning("No activities found in this file.")
        return

    tasks = get_critical_threshold(tasks, near_crit_days)

    # Header metrics
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Project Name", proj.get("name", "Unknown"))
    c2.metric("Data Date", format_date(proj.get("data_date")))
    c3.metric("Parse Method", data.get("parse_method", "-"))
    c4.metric("Activities", len(tasks))

    st.divider()
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("🔴 Critical", int(tasks["is_critical"].sum()))
    c2.metric("🟡 Near-Critical", int(tasks["is_near_critical"].sum()))
    neg_float = tasks["total_float_days"].apply(lambda f: f is not None and f < 0).sum()
    c3.metric("⚠️ Negative Float", int(neg_float))
    c4.metric("🔗 Relationships", len(rels))

    # Open-ended
    if not rels.empty and "pred_task_id" in rels.columns:
        tasks_with_pred = set(rels["succ_task_id"].dropna())
        tasks_with_succ = set(rels["pred_task_id"].dropna())
        task_ids = set(tasks["task_id"])
        no_pred = len(task_ids - tasks_with_pred)
        no_succ = len(task_ids - tasks_with_succ)
        c5.metric("Open-Ended Activities", no_pred + no_succ)
    else:
        c5.metric("Open-Ended Activities", "-")

    # Date range
    valid_starts = tasks["eff_start"].dropna()
    valid_finishes = tasks["eff_finish"].dropna()
    if not valid_starts.empty and not valid_finishes.empty:
        earliest = min(valid_starts)
        latest = max(valid_finishes)
        st.info(f"**Schedule Span:** {format_date(earliest)} -> {format_date(latest)}")

    # Constraint count
    constrained = tasks["cstr_type"].apply(lambda x: bool(x) and str(x).strip() not in ("", "None")).sum() if "cstr_type" in tasks.columns else 0
    st.info(f"**Constrained Activities:** {int(constrained)}")

    st.divider()

    # Charts
    tab1, tab2, tab3 = st.tabs(["Float Distribution", "Activities by WBS", "Status Breakdown"])

    with tab1:
        float_vals = tasks["total_float_days"].dropna()
        if not float_vals.empty:
            fig = px.histogram(
                float_vals, nbins=40, title="Total Float Distribution (Days)",
                labels={"value": "Float (days)", "count": "Activities"},
                color_discrete_sequence=["#2563eb"],
            )
            fig.add_vline(x=0, line_dash="dash", line_color="red", annotation_text="Critical")
            fig.add_vline(x=near_crit_days, line_dash="dot", line_color="orange",
                          annotation_text=f"Near-Critical ({near_crit_days}d)")
            st.plotly_chart(fig, use_container_width=True)

    with tab2:
        if "wbs_path" in tasks.columns:
            # Show top-level WBS only
            tasks["wbs_top"] = tasks["wbs_path"].apply(
                lambda x: str(x).split(" > ")[0] if pd.notna(x) and x else "Unknown"
            )
            wbs_counts = tasks.groupby("wbs_top").size().reset_index(name="count")
            wbs_counts = wbs_counts.sort_values("count", ascending=False).head(20)
            fig = px.bar(wbs_counts, x="count", y="wbs_top", orientation="h",
                         title="Activities by Top-Level WBS",
                         color_discrete_sequence=["#1e3a5f"])
            fig.update_layout(yaxis_title="", xaxis_title="Activity Count")
            st.plotly_chart(fig, use_container_width=True)

    with tab3:
        if "status" in tasks.columns:
            status_counts = tasks["status"].value_counts().reset_index()
            status_counts.columns = ["Status", "Count"]
            fig = px.pie(status_counts, values="Count", names="Status",
                         title="Activity Status Breakdown",
                         color_discrete_sequence=px.colors.qualitative.Set2)
            st.plotly_chart(fig, use_container_width=True)

    # Summary table
    st.subheader("Activity Summary Table")
    display_cols = ["task_code", "task_name", "wbs_path", "eff_start", "eff_finish",
                    "total_float_days", "status", "is_critical"]
    avail = [c for c in display_cols if c in tasks.columns]
    st.dataframe(tasks[avail].head(100), use_container_width=True)


# -----------------------------------------------------------------------------
# PAGE: ACTIVITY SEARCH
# -----------------------------------------------------------------------------

def _col(df: pd.DataFrame, col: str, default="-"):
    """Safely get a column value; return default if column missing or null."""
    if col not in df.index:
        return default
    val = df.get(col, default)
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return default
    return val


def _float_color(f) -> str:
    """Return a hex colour string for a float value."""
    if f is None:
        return "#6b7280"
    try:
        f = float(f)
    except (TypeError, ValueError):
        return "#6b7280"
    if f < 0:
        return "#991b1b"   # dark red  - negative float
    if f == 0:
        return "#dc2626"   # red       - critical
    if f <= 10:
        return "#d97706"   # amber     - near-critical
    return "#15803d"       # green     - has float


def _status_label(status: str) -> str:
    """Convert P6 status code to a readable label."""
    mapping = {
        "TK_NotStart": "Not Started",
        "TK_Active":   "In Progress",
        "TK_Complete": "Complete",
        "Not Started": "Not Started",
        "In Progress": "In Progress",
        "Complete":    "Complete",
    }
    return mapping.get(str(status).strip(), str(status).strip() or "-")


def _status_colour(status: str) -> str:
    s = _status_label(status)
    if s == "Complete":    return "#15803d"
    if s == "In Progress": return "#2563eb"
    return "#6b7280"


def page_activity_search(data: dict, near_crit_days: float):
    """
    Activity Search page.
    Lets the user filter activities by ID, name, WBS, date range and
    critical status, then shows a full detail panel for the selected activity.
    """
    st.title("🔍 Activity Search")
    st.caption("Search and filter activities, then click any row to view its full detail.")

    tasks = data["tasks_df"]
    rels  = data["relationships_df"]

    if tasks.empty:
        st.warning("No activities found. Please upload an XER file first.")
        return

    tasks = get_critical_threshold(tasks, near_crit_days)

    # -------------------------------------------------------------------------
    # SEARCH / FILTER PANEL
    # -------------------------------------------------------------------------
    with st.expander("🔎  Search & Filter", expanded=True):
        r1c1, r1c2, r1c3 = st.columns(3)
        search_code = r1c1.text_input("Activity ID", placeholder="e.g. A1000")
        search_name = r1c2.text_input("Activity Name", placeholder="partial match")
        search_wbs  = r1c3.text_input("WBS", placeholder="partial match")

        r2c1, r2c2, r2c3 = st.columns(3)

        # Critical / float filter
        crit_filter = r2c1.selectbox(
            "Float Status",
            ["All", "Critical (float <= 0)", "Near-Critical", "Positive Float", "Negative Float"],
        )

        # Status filter - only show if column exists
        if "status" in tasks.columns:
            status_opts = ["All"] + sorted(
                [s for s in tasks["status"].dropna().unique() if str(s).strip()]
            )
        else:
            status_opts = ["All"]
        status_filter = r2c2.selectbox("Activity Status", status_opts)

        # Activity type filter
        if "task_type" in tasks.columns:
            type_opts = ["All"] + sorted(
                [t for t in tasks["task_type"].dropna().unique() if str(t).strip()]
            )
        else:
            type_opts = ["All"]
        type_filter = r2c3.selectbox("Activity Type", type_opts)

        # Date range - only show if date columns exist and have data
        date_from = date_to = None
        valid_starts  = tasks["eff_start"].dropna()  if "eff_start"  in tasks.columns else pd.Series(dtype=object)
        valid_finishes = tasks["eff_finish"].dropna() if "eff_finish" in tasks.columns else pd.Series(dtype=object)

        if not valid_starts.empty and not valid_finishes.empty:
            try:
                min_d = min(valid_starts).date()
                max_d = max(valid_finishes).date()
                dc1, dc2 = st.columns(2)
                date_from = dc1.date_input("Finish on or After", value=min_d,
                                           min_value=min_d, max_value=max_d)
                date_to   = dc2.date_input("Finish on or Before", value=max_d,
                                           min_value=min_d, max_value=max_d)
            except Exception:
                pass  # silently skip date filter if dates are unusable

    # -------------------------------------------------------------------------
    # APPLY FILTERS
    # -------------------------------------------------------------------------
    filtered = tasks.copy()

    if search_code.strip():
        if "task_code" in filtered.columns:
            filtered = filtered[
                filtered["task_code"].astype(str).str.contains(
                    search_code.strip(), case=False, na=False
                )
            ]
    if search_name.strip():
        if "task_name" in filtered.columns:
            filtered = filtered[
                filtered["task_name"].astype(str).str.contains(
                    search_name.strip(), case=False, na=False
                )
            ]
    if search_wbs.strip():
        if "wbs_path" in filtered.columns:
            filtered = filtered[
                filtered["wbs_path"].astype(str).str.contains(
                    search_wbs.strip(), case=False, na=False
                )
            ]

    if crit_filter == "Critical (float <= 0)" and "is_critical" in filtered.columns:
        filtered = filtered[filtered["is_critical"] == True]
    elif crit_filter == "Near-Critical" and "is_near_critical" in filtered.columns:
        filtered = filtered[filtered["is_near_critical"] == True]
    elif crit_filter == "Positive Float" and "total_float_days" in filtered.columns:
        filtered = filtered[filtered["total_float_days"].apply(
            lambda f: f is not None and safe_float(f, 1) > 0
        )]
    elif crit_filter == "Negative Float" and "total_float_days" in filtered.columns:
        filtered = filtered[filtered["total_float_days"].apply(
            lambda f: f is not None and safe_float(f, 0) < 0
        )]

    if status_filter != "All" and "status" in filtered.columns:
        filtered = filtered[filtered["status"] == status_filter]

    if type_filter != "All" and "task_type" in filtered.columns:
        filtered = filtered[filtered["task_type"] == type_filter]

    if date_from is not None and "eff_finish" in filtered.columns:
        filtered = filtered[
            filtered["eff_finish"].apply(
                lambda d: d is not None and hasattr(d, "date") and d.date() >= date_from
            )
        ]
    if date_to is not None and "eff_finish" in filtered.columns:
        filtered = filtered[
            filtered["eff_finish"].apply(
                lambda d: d is not None and hasattr(d, "date") and d.date() <= date_to
            )
        ]

    # -------------------------------------------------------------------------
    # RESULTS TABLE
    # -------------------------------------------------------------------------
    n_found = len(filtered)
    n_total = len(tasks)

    if n_found == 0:
        st.warning("No activities match your filters. Try broadening your search.")
        return

    # Build a clean display version of the table
    TABLE_COLS = {
        "task_code":        "Activity ID",
        "task_name":        "Activity Name",
        "wbs_path":         "WBS",
        "eff_start":        "Start",
        "eff_finish":       "Finish",
        "orig_dur_days":    "Orig Dur (d)",
        "rem_dur_days":     "Rem Dur (d)",
        "total_float_days": "Float (d)",
        "status":           "Status",
        "task_type":        "Type",
        "is_critical":      "Critical",
    }

    present_cols = {k: v for k, v in TABLE_COLS.items() if k in filtered.columns}
    display_df = filtered[list(present_cols.keys())].copy()
    display_df = display_df.rename(columns=present_cols)

    # Format date columns for readability
    for col in ["Start", "Finish"]:
        if col in display_df.columns:
            display_df[col] = display_df[col].apply(format_date)

    # Format critical flag
    if "Critical" in display_df.columns:
        display_df["Critical"] = display_df["Critical"].apply(
            lambda x: "Yes" if x else ""
        )

    # Friendly status labels
    if "Status" in display_df.columns:
        display_df["Status"] = display_df["Status"].apply(_status_label)

    st.markdown(
        f"<p style='color:#6b7280;font-size:13px;'>"
        f"Showing <strong>{n_found}</strong> of <strong>{n_total}</strong> activities"
        f"</p>",
        unsafe_allow_html=True,
    )

    st.dataframe(
        display_df,
        use_container_width=True,
        height=min(400, 45 + n_found * 35),
        hide_index=True,
    )

    # -------------------------------------------------------------------------
    # ACTIVITY SELECTOR  -  pick from the filtered results
    # -------------------------------------------------------------------------
    st.divider()

    # Build selector labels
    def make_label(r):
        code = str(r.get("task_code", "?"))
        name = str(r.get("task_name", "?"))
        return f"{code}  —  {name}"

    act_labels = filtered.apply(make_label, axis=1).tolist()

    selected_label = st.selectbox(
        "Select an activity to view full details",
        options=act_labels,
        index=0,
    )

    sel_idx = act_labels.index(selected_label)
    row = filtered.iloc[sel_idx]

    # -------------------------------------------------------------------------
    # SELECTED ACTIVITY BANNER
    # -------------------------------------------------------------------------
    tf_val  = row.get("total_float_days") if "total_float_days" in row.index else None
    tf_num  = safe_float(tf_val, None)
    f_color = _float_color(tf_num)

    status_raw = row.get("status", "") if "status" in row.index else ""
    s_label    = _status_label(str(status_raw))
    s_color    = _status_colour(str(status_raw))

    is_crit = bool(row.get("is_critical", False)) if "is_critical" in row.index else False
    crit_banner = (
        '<span style="background:#dc2626;color:white;padding:3px 10px;'
        'border-radius:12px;font-size:12px;font-weight:700;margin-left:10px;">CRITICAL</span>'
        if is_crit else ""
    )

    task_code = str(row.get("task_code", "-")) if "task_code" in row.index else "-"
    task_name = str(row.get("task_name", "-")) if "task_name" in row.index else "-"

    st.markdown(
        f"""
        <div style="background:#1e3a5f;color:white;border-radius:10px;
                    padding:18px 24px;margin-bottom:16px;">
            <div style="font-size:13px;color:#93c5fd;font-weight:600;
                        letter-spacing:1px;text-transform:uppercase;">
                Selected Activity
            </div>
            <div style="font-size:22px;font-weight:700;margin-top:4px;">
                {task_code}{crit_banner}
            </div>
            <div style="font-size:16px;color:#bfdbfe;margin-top:4px;">
                {task_name}
            </div>
            <div style="margin-top:10px;">
                <span style="background:{s_color};color:white;padding:3px 10px;
                             border-radius:12px;font-size:12px;font-weight:600;">
                    {s_label}
                </span>
                <span style="background:{f_color};color:white;padding:3px 10px;
                             border-radius:12px;font-size:12px;font-weight:600;
                             margin-left:8px;">
                    Float: {tf_num if tf_num is not None else "-"} days
                </span>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # -------------------------------------------------------------------------
    # DETAIL PANEL  -  two column layout
    # -------------------------------------------------------------------------
    def field(label: str, value, suffix: str = "") -> str:
        """Render a labelled field as HTML. Handles missing/None gracefully."""
        if value is None or str(value).strip() in ("", "None", "nan", "-"):
            disp = '<span style="color:#9ca3af;">Not available</span>'
        else:
            disp = f'<span style="font-weight:600;color:#111827;">{value}{suffix}</span>'
        return (
            f'<div style="padding:6px 0;border-bottom:1px solid #f3f4f6;">'
            f'<span style="color:#6b7280;font-size:12px;text-transform:uppercase;'
            f'letter-spacing:0.5px;">{label}</span><br>{disp}</div>'
        )

    col_left, col_right = st.columns(2)

    with col_left:
        st.markdown("##### Identity & Classification")
        st.markdown(
            field("Activity ID",    _col(row, "task_code")) +
            field("Activity Name",  _col(row, "task_name")) +
            field("WBS",            _col(row, "wbs_path")) +
            field("Activity Type",  _col(row, "task_type")) +
            field("Calendar",       _col(row, "calendar")) +
            field("% Complete",     _col(row, "phys_pct"), suffix="%"),
            unsafe_allow_html=True,
        )

    with col_right:
        st.markdown("##### Schedule & Float")
        early_start  = format_date(row.get("early_start")  if "early_start"  in row.index else None)
        early_finish = format_date(row.get("early_finish") if "early_finish" in row.index else None)
        late_start   = format_date(row.get("late_start")   if "late_start"   in row.index else None)
        late_finish  = format_date(row.get("late_finish")  if "late_finish"  in row.index else None)
        act_start    = format_date(row.get("act_start")    if "act_start"    in row.index else None)
        act_finish   = format_date(row.get("act_finish")   if "act_finish"   in row.index else None)
        orig_dur     = _col(row, "orig_dur_days")
        rem_dur      = _col(row, "rem_dur_days")
        ff_val       = _col(row, "free_float_days")

        st.markdown(
            field("Early Start",         early_start) +
            field("Early Finish",        early_finish) +
            field("Late Start",          late_start) +
            field("Late Finish",         late_finish) +
            field("Actual Start",        act_start) +
            field("Actual Finish",       act_finish) +
            field("Original Duration",   orig_dur,  suffix=" days") +
            field("Remaining Duration",  rem_dur,   suffix=" days") +
            field("Total Float",         tf_num,    suffix=" days") +
            field("Free Float",          ff_val,    suffix=" days"),
            unsafe_allow_html=True,
        )

    # Constraint row - only show when a constraint is set
    if "cstr_type" in row.index:
        cstr = row.get("cstr_type", "")
        if cstr and str(cstr).strip() not in ("", "None", "nan"):
            cstr_date = format_date(row.get("cstr_date") if "cstr_date" in row.index else None)
            st.markdown(
                f'<div style="background:#fef3c7;border-left:4px solid #f59e0b;'
                f'border-radius:6px;padding:10px 16px;margin-top:12px;">'
                f'<strong>Constraint:</strong> {cstr} &nbsp;|&nbsp; '
                f'<strong>Constraint Date:</strong> {cstr_date}</div>',
                unsafe_allow_html=True,
            )

    # -------------------------------------------------------------------------
    # PREDECESSORS & SUCCESSORS
    # -------------------------------------------------------------------------
    st.markdown("---")
    st.markdown("##### Logic")

    pred_col, succ_col = st.columns(2)

    if not rels.empty and "task_id" in row.index:
        task_id = row["task_id"]

        # --- Predecessors ---
        with pred_col:
            st.markdown("**Predecessors**  *(what drives this activity)*")
            if "succ_task_id" in rels.columns:
                preds = rels[rels["succ_task_id"] == task_id].copy()
            else:
                preds = pd.DataFrame()

            if preds.empty:
                st.info("No predecessors — this is an open start.")
            else:
                pred_display_cols = {
                    "pred_task_code": "Activity ID",
                    "pred_task_name": "Activity Name",
                    "rel_type":       "Link",
                    "lag_days":       "Lag (d)",
                }
                pred_show = {k: v for k, v in pred_display_cols.items()
                             if k in preds.columns}
                pred_df = preds[list(pred_show.keys())].rename(columns=pred_show)
                st.dataframe(pred_df, use_container_width=True, hide_index=True)

        # --- Successors ---
        with succ_col:
            st.markdown("**Successors**  *(what this activity drives)*")
            if "pred_task_id" in rels.columns:
                succs = rels[rels["pred_task_id"] == task_id].copy()
            else:
                succs = pd.DataFrame()

            if succs.empty:
                st.info("No successors — this is an open finish.")
            else:
                succ_display_cols = {
                    "succ_task_code": "Activity ID",
                    "succ_task_name": "Activity Name",
                    "rel_type":       "Link",
                    "lag_days":       "Lag (d)",
                }
                succ_show = {k: v for k, v in succ_display_cols.items()
                             if k in succs.columns}
                succ_df = succs[list(succ_show.keys())].rename(columns=succ_show)
                st.dataframe(succ_df, use_container_width=True, hide_index=True)
    else:
        with pred_col:
            st.info("Relationship data not available.")
        with succ_col:
            st.info("Relationship data not available.")

    # -------------------------------------------------------------------------
    # EXPORT THIS ACTIVITY
    # -------------------------------------------------------------------------
    st.markdown("---")

    # Build a single-row detail export
    export_fields = {
        "Activity ID":        _col(row, "task_code"),
        "Activity Name":      _col(row, "task_name"),
        "WBS":                _col(row, "wbs_path"),
        "Activity Type":      _col(row, "task_type"),
        "Calendar":           _col(row, "calendar"),
        "Status":             _status_label(str(_col(row, "status", ""))),
        "% Complete":         _col(row, "phys_pct"),
        "Early Start":        early_start,
        "Early Finish":       early_finish,
        "Late Start":         late_start,
        "Late Finish":        late_finish,
        "Actual Start":       act_start,
        "Actual Finish":      act_finish,
        "Original Duration":  _col(row, "orig_dur_days"),
        "Remaining Duration": _col(row, "rem_dur_days"),
        "Total Float (days)": tf_num,
        "Free Float (days)":  _col(row, "free_float_days"),
        "Critical":           "Yes" if is_crit else "No",
        "Constraint":         _col(row, "cstr_type"),
        "Constraint Date":    format_date(row.get("cstr_date") if "cstr_date" in row.index else None),
    }
    detail_df = pd.DataFrame([export_fields])

    # Predecessors / successors for export
    export_sheets = {"Activity Detail": detail_df}

    if not rels.empty and "task_id" in row.index:
        task_id = row["task_id"]
        if not preds.empty:
            pred_exp_cols = [c for c in ["pred_task_code","pred_task_name","rel_type","lag_days"]
                             if c in preds.columns]
            export_sheets["Predecessors"] = preds[pred_exp_cols].rename(columns={
                "pred_task_code": "Activity ID", "pred_task_name": "Activity Name",
                "rel_type": "Link", "lag_days": "Lag (days)"
            })
        if not succs.empty:
            succ_exp_cols = [c for c in ["succ_task_code","succ_task_name","rel_type","lag_days"]
                             if c in succs.columns]
            export_sheets["Successors"] = succs[succ_exp_cols].rename(columns={
                "succ_task_code": "Activity ID", "succ_task_name": "Activity Name",
                "rel_type": "Link", "lag_days": "Lag (days)"
            })

    xls_bytes = export_df_to_excel(export_sheets)
    st.download_button(
        label="📥  Export Activity Detail to Excel",
        data=xls_bytes,
        file_name=f"activity_{_col(row, 'task_code', 'detail')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# -----------------------------------------------------------------------------
# PAGE: LOGIC TRACE
# -----------------------------------------------------------------------------

# Relationship type codes -> readable labels
REL_TYPE_LABELS = {
    "FS": "Finish-to-Start (FS)",
    "FF": "Finish-to-Finish (FF)",
    "SS": "Start-to-Start (SS)",
    "SF": "Start-to-Finish (SF)",
    "PR_FS": "Finish-to-Start (FS)",
    "PR_FF": "Finish-to-Finish (FF)",
    "PR_SS": "Start-to-Start (SS)",
    "PR_SF": "Start-to-Finish (SF)",
}


def _rel_label(code: str) -> str:
    return REL_TYPE_LABELS.get(str(code).strip(), str(code).strip() or "-")


def _crit_flag(tf) -> str:
    """Return a text critical flag suitable for display and export."""
    if tf is None:
        return "-"
    try:
        f = float(tf)
    except (TypeError, ValueError):
        return "-"
    if f < 0:
        return "Negative Float"
    if f == 0:
        return "Critical"
    if f <= 10:
        return "Near-Critical"
    return "Float"


def _build_full_trace_df(
    G: nx.DiGraph,
    rels_df: pd.DataFrame,
    task_lookup: dict,
    selected_id: str,
    trace_list: list,     # [(task_id, depth), ...]
    direction: str,       # "pred" | "succ" | "both"
) -> pd.DataFrame:
    """
    Build the trace results DataFrame.

    For each activity in trace_list, look up the relationship that
    connects it to the selected activity (or to the activity one step
    closer in the chain) so we can display the correct link type and lag.

    direction:
        "pred"  - activities are predecessors; depth counts backwards  (1 = direct pred)
        "succ"  - activities are successors;   depth counts forwards   (1 = direct succ)
        "both"  - mixed: negative depth = pred side, positive = succ side
    """
    rows = []

    for tid, depth in trace_list:
        t = task_lookup.get(tid, {})
        tf = t.get("total_float_days")

        # ---- find the relationship between this activity and the one at
        #      depth-1 in the chain (i.e. the step that led here) ----------
        rel_type_str = "-"
        lag_val = 0

        if not rels_df.empty:
            if direction in ("pred", "both") and depth > 0:
                # this activity is a predecessor of something; find its
                # outgoing relationship toward the selected direction
                mask = (
                    (rels_df.get("pred_task_id", pd.Series(dtype=str)) == tid) |
                    (rels_df.get("succ_task_id", pd.Series(dtype=str)) == tid)
                )
                rel_candidates = rels_df[mask]
            elif direction == "succ":
                mask = (
                    (rels_df.get("pred_task_id", pd.Series(dtype=str)) == tid) |
                    (rels_df.get("succ_task_id", pd.Series(dtype=str)) == tid)
                )
                rel_candidates = rels_df[mask]
            else:
                rel_candidates = pd.DataFrame()

            if not rel_candidates.empty and "rel_type" in rel_candidates.columns:
                rel_type_str = _rel_label(rel_candidates["rel_type"].iloc[0])
                lag_val = rel_candidates["lag_days"].iloc[0] if "lag_days" in rel_candidates.columns else 0
                try:
                    lag_val = float(lag_val) if lag_val is not None else 0
                except (TypeError, ValueError):
                    lag_val = 0

        rows.append({
            "Depth": depth,
            "Direction": "Predecessor" if depth < 0 or direction == "pred"
                         else ("Successor" if direction == "succ" else "Both"),
            "Activity ID":    t.get("task_code", tid),
            "Activity Name":  t.get("task_name", ""),
            "Link Type":      rel_type_str,
            "Lag (days)":     lag_val,
            "Start":          format_date(t.get("eff_start")),
            "Finish":         format_date(t.get("eff_finish")),
            "Total Float (d)": tf if tf is not None else "-",
            "Status":         _status_label(str(t.get("status", ""))),
            "Critical Flag":  _crit_flag(tf),
        })

    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values("Depth").reset_index(drop=True)
    return df


def _summary_bar(label: str, value: int, colour: str) -> str:
    return (
        f'<div style="display:inline-block;background:{colour};color:white;'
        f'border-radius:8px;padding:8px 18px;margin:4px 6px 4px 0;font-size:13px;">'
        f'<strong>{value}</strong> {label}</div>'
    )


def page_logic_trace(data: dict, near_crit_days: float):
    """
    Logic Trace page.
    Select any activity and use the buttons to trace direct or full
    predecessor/successor chains through the schedule network.
    Each result row shows depth, link type, lag, dates, float and
    critical status. Results persist while the same activity is selected.
    """
    st.title("🔗 Logic Trace")
    st.caption(
        "Select an activity then use the buttons below to trace its logic. "
        "Activities are treated as network nodes - relationships are the edges connecting them."
    )

    tasks = data["tasks_df"]
    rels  = data["relationships_df"]

    # -------------------------------------------------------------------------
    # GUARD RAILS
    # -------------------------------------------------------------------------
    if tasks.empty:
        st.warning("No activities found. Please upload an XER file first.")
        return

    tasks = get_critical_threshold(tasks, near_crit_days)

    rels_available = not rels.empty
    if not rels_available:
        st.warning(
            "No relationship data was found in this XER file.  "
            "Logic tracing requires both activities and relationships to be present."
        )
        # Still show the activity list so the user can at least see the data loaded
        st.dataframe(
            tasks[["task_code", "task_name", "total_float_days", "status"]].head(50),
            use_container_width=True, hide_index=True,
        )
        return

    # -------------------------------------------------------------------------
    # BUILD NETWORK GRAPH
    # -------------------------------------------------------------------------
    G           = build_graph(tasks, rels)
    task_lookup = tasks.set_index("task_id").to_dict("index")

    # -------------------------------------------------------------------------
    # ACTIVITY SELECTOR
    # -------------------------------------------------------------------------
    def _act_label(r):
        return f"{r.get('task_code','?')}  --  {r.get('task_name','?')}"

    act_labels = tasks.apply(_act_label, axis=1).tolist()

    selected_label = st.selectbox(
        "Select activity to trace",
        options=act_labels,
        key="logic_trace_selector",
    )
    sel_idx      = act_labels.index(selected_label)
    selected_row = tasks.iloc[sel_idx]
    selected_id  = selected_row["task_id"]
    sel_code     = str(selected_row.get("task_code", "-"))
    sel_name     = str(selected_row.get("task_name", "-"))

    # Clear previous results if the user picks a different activity
    if st.session_state.get("_trace_last_id") != selected_id:
        for k in ("trace_df", "trace_label", "trace_direction"):
            st.session_state.pop(k, None)
        st.session_state["_trace_last_id"] = selected_id

    # -------------------------------------------------------------------------
    # SELECTED ACTIVITY BANNER
    # -------------------------------------------------------------------------
    sel_tf    = safe_float(selected_row.get("total_float_days"), None) if "total_float_days" in selected_row.index else None
    sel_fcol  = _float_color(sel_tf)
    sel_crit  = bool(selected_row.get("is_critical", False)) if "is_critical" in selected_row.index else False
    sel_stat  = _status_label(str(selected_row.get("status", "")))
    sel_scol  = _status_colour(str(selected_row.get("status", "")))
    crit_pill = (
        '<span style="background:#dc2626;color:white;padding:2px 10px;'
        'border-radius:12px;font-size:11px;font-weight:700;margin-left:8px;">CRITICAL</span>'
        if sel_crit else ""
    )

    st.markdown(
        f"""
        <div style="background:#1e3a5f;color:white;border-radius:10px;
                    padding:16px 22px;margin:8px 0 18px 0;">
            <div style="font-size:12px;color:#93c5fd;font-weight:600;
                        letter-spacing:1px;text-transform:uppercase;">
                Tracing Activity
            </div>
            <div style="font-size:20px;font-weight:700;margin-top:4px;">
                {sel_code}{crit_pill}
            </div>
            <div style="font-size:14px;color:#bfdbfe;margin-top:2px;">{sel_name}</div>
            <div style="margin-top:10px;">
                <span style="background:{sel_scol};color:white;padding:3px 10px;
                             border-radius:12px;font-size:12px;">{sel_stat}</span>
                <span style="background:{sel_fcol};color:white;padding:3px 10px;
                             border-radius:12px;font-size:12px;margin-left:6px;">
                    Float: {sel_tf if sel_tf is not None else "-"} days
                </span>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # -------------------------------------------------------------------------
    # OPEN-END WARNINGS  (check before showing buttons)
    # -------------------------------------------------------------------------
    direct_preds = list(G.predecessors(selected_id))
    direct_succs = list(G.successors(selected_id))
    has_preds    = len(direct_preds) > 0
    has_succs    = len(direct_succs) > 0

    if not has_preds:
        st.warning(
            f"**Open Start:** {sel_code} has no predecessors. "
            "This activity is open-ended at the start - it is not driven by any logic in the programme."
        )
    if not has_succs:
        st.warning(
            f"**Open Finish:** {sel_code} has no successors. "
            "This activity is open-ended at the finish - no activities are driven by it in the programme."
        )

    if not has_preds and not has_succs:
        st.error(
            "This activity has no logic connections at all. "
            "It is completely isolated in the programme network."
        )
        return

    # -------------------------------------------------------------------------
    # QUICK STATS ROW  (counts before any buttons pressed)
    # -------------------------------------------------------------------------
    all_pred_ids = [tid for tid, _ in trace_predecessors(G, selected_id)]
    all_succ_ids = [tid for tid, _ in trace_successors(G, selected_id)]

    n_dp  = len(direct_preds)
    n_ds  = len(direct_succs)
    n_ap  = len(all_pred_ids)
    n_as  = len(all_succ_ids)

    # How many in the full pred/succ network are critical?
    def count_crit(id_list):
        return sum(
            1 for tid in id_list
            if safe_float(task_lookup.get(tid, {}).get("total_float_days"), 1) <= 0
        )

    st.markdown(
        _summary_bar(f"Direct predecessors", n_dp, "#1d4ed8") +
        _summary_bar(f"Direct successors",   n_ds, "#1d4ed8") +
        _summary_bar(f"All predecessors",    n_ap, "#4338ca") +
        _summary_bar(f"All successors",      n_as, "#4338ca") +
        _summary_bar(f"Critical in pred network", count_crit(all_pred_ids), "#dc2626") +
        _summary_bar(f"Critical in succ network", count_crit(all_succ_ids), "#dc2626"),
        unsafe_allow_html=True,
    )

    st.markdown("<br>", unsafe_allow_html=True)

    # -------------------------------------------------------------------------
    # ACTION BUTTONS
    # -------------------------------------------------------------------------
    st.markdown("**Choose what to trace:**")

    b1, b2, b3, b4, b5 = st.columns(5)
    btn_dir_pred  = b1.button("◀  Direct\nPredecessors",  key="btn_dp",  use_container_width=True, disabled=not has_preds)
    btn_dir_succ  = b2.button("▶  Direct\nSuccessors",    key="btn_ds",  use_container_width=True, disabled=not has_succs)
    btn_all_pred  = b3.button("◀◀  All\nPredecessors",    key="btn_ap",  use_container_width=True, disabled=not has_preds)
    btn_all_succ  = b4.button("▶▶  All\nSuccessors",      key="btn_as",  use_container_width=True, disabled=not has_succs)
    btn_full      = b5.button("Full Logic\nChain",         key="btn_fl",  use_container_width=True)

    # -------------------------------------------------------------------------
    # HANDLE BUTTON PRESSES  ->  build trace list
    # -------------------------------------------------------------------------
    new_trace  = None   # list of (task_id, depth)
    new_label  = None
    new_dir    = None

    if btn_dir_pred:
        new_trace = [(p, 1) for p in direct_preds]
        new_label = f"Direct Predecessors of {sel_code}"
        new_dir   = "pred"

    elif btn_dir_succ:
        new_trace = [(s, 1) for s in direct_succs]
        new_label = f"Direct Successors of {sel_code}"
        new_dir   = "succ"

    elif btn_all_pred:
        new_trace = trace_predecessors(G, selected_id)
        new_label = f"All Predecessors of {sel_code}"
        new_dir   = "pred"

    elif btn_all_succ:
        new_trace = trace_successors(G, selected_id)
        new_label = f"All Successors of {sel_code}"
        new_dir   = "succ"

    elif btn_full:
        preds_list = [(tid, -depth) for tid, depth in trace_predecessors(G, selected_id)]
        succs_list = [(tid,  depth) for tid, depth in trace_successors(G,  selected_id)]
        new_trace  = preds_list + succs_list
        new_label  = f"Full Logic Chain for {sel_code}"
        new_dir    = "both"

    # Persist results across reruns
    if new_trace is not None:
        trace_df = _build_full_trace_df(
            G, rels, task_lookup, selected_id, new_trace, new_dir
        )
        st.session_state["trace_df"]        = trace_df
        st.session_state["trace_label"]     = new_label
        st.session_state["trace_direction"] = new_dir

    # -------------------------------------------------------------------------
    # RESULTS DISPLAY
    # -------------------------------------------------------------------------
    if "trace_df" in st.session_state and not st.session_state["trace_df"].empty:
        trace_df  = st.session_state["trace_df"]
        trace_lbl = st.session_state.get("trace_label", "Trace Results")
        trace_dir = st.session_state.get("trace_direction", "pred")

        st.divider()

        # Result header
        n_res   = len(trace_df)
        n_crit  = int((trace_df["Critical Flag"] == "Critical").sum())
        n_neg   = int((trace_df["Critical Flag"] == "Negative Float").sum())
        n_near  = int((trace_df["Critical Flag"] == "Near-Critical").sum())

        st.markdown(
            f"<h4 style='color:#1e3a5f;margin-bottom:4px;'>{trace_lbl}</h4>",
            unsafe_allow_html=True,
        )
        st.markdown(
            _summary_bar(f"activities", n_res, "#374151") +
            (_summary_bar("critical",      n_crit, "#dc2626") if n_crit else "") +
            (_summary_bar("negative float",n_neg,  "#7f1d1d") if n_neg  else "") +
            (_summary_bar("near-critical", n_near, "#d97706") if n_near else ""),
            unsafe_allow_html=True,
        )
        st.markdown("<br>", unsafe_allow_html=True)

        # Direction legend
        if trace_dir == "pred":
            st.caption(
                "Depth = how many steps back this activity is from your selected activity. "
                "Depth 1 = directly connected."
            )
        elif trace_dir == "succ":
            st.caption(
                "Depth = how many steps forward this activity is from your selected activity. "
                "Depth 1 = directly connected."
            )
        else:
            st.caption(
                "Negative depth = predecessors (how many steps back). "
                "Positive depth = successors (how many steps forward)."
            )

        # Colour-code the Critical Flag column in the dataframe
        def _colour_flag(val):
            colour_map = {
                "Critical":       "background-color:#fee2e2;color:#991b1b;font-weight:600;",
                "Negative Float": "background-color:#fecaca;color:#7f1d1d;font-weight:700;",
                "Near-Critical":  "background-color:#fef3c7;color:#92400e;font-weight:600;",
                "Float":          "background-color:#dcfce7;color:#166534;",
            }
            return colour_map.get(val, "")

        styled = trace_df.style.applymap(_colour_flag, subset=["Critical Flag"])

        st.dataframe(styled, use_container_width=True, hide_index=True, height=min(600, 45 + n_res * 35))

        # ---- TABS for filtered views ----------------------------------------
        if n_res > 5:
            st.markdown("**Filter results:**")
            tab_all, tab_crit, tab_near, tab_open = st.tabs(
                ["All", "Critical & Negative Float", "Near-Critical", "Open Ends"]
            )

            with tab_all:
                st.dataframe(trace_df, use_container_width=True, hide_index=True)

            with tab_crit:
                crit_df = trace_df[trace_df["Critical Flag"].isin(["Critical", "Negative Float"])]
                if crit_df.empty:
                    st.success("No critical activities in this trace.")
                else:
                    st.dataframe(crit_df, use_container_width=True, hide_index=True)

            with tab_near:
                near_df = trace_df[trace_df["Critical Flag"] == "Near-Critical"]
                if near_df.empty:
                    st.success("No near-critical activities in this trace.")
                else:
                    st.dataframe(near_df, use_container_width=True, hide_index=True)

            with tab_open:
                # Activities in the trace that themselves have open ends
                open_rows = []
                for tid in trace_df["Activity ID"].tolist():
                    # Map display code back to task_id
                    match = tasks[tasks["task_code"] == tid]
                    if match.empty:
                        continue
                    mid = match.iloc[0]["task_id"]
                    no_p = len(list(G.predecessors(mid))) == 0
                    no_s = len(list(G.successors(mid))) == 0
                    if no_p or no_s:
                        t = task_lookup.get(mid, {})
                        open_rows.append({
                            "Activity ID":   t.get("task_code", mid),
                            "Activity Name": t.get("task_name", ""),
                            "Issue":         ("No predecessors" if no_p else "") +
                                             (" | No successors" if no_s else ""),
                        })
                if open_rows:
                    st.dataframe(pd.DataFrame(open_rows), use_container_width=True, hide_index=True)
                else:
                    st.success("No open-ended activities found in this trace.")

        # ---- GANTT chart for the trace results ------------------------------
        st.divider()
        st.markdown("**Trace Timeline**")

        # Merge trace_df back with tasks to get proper date objects
        gantt_data = trace_df.copy()
        gantt_data = gantt_data.merge(
            tasks[["task_code", "eff_start", "eff_finish", "is_critical"]].rename(
                columns={"task_code": "Activity ID"}
            ),
            on="Activity ID",
            how="left",
        )
        gantt_data = gantt_data.dropna(subset=["eff_start", "eff_finish"])

        if not gantt_data.empty:
            gantt_data["Label"] = gantt_data["Activity ID"] + "  " + gantt_data["Activity Name"].str[:40]
            gantt_data["Bar Colour"] = gantt_data["Critical Flag"].map({
                "Critical":       "Critical",
                "Negative Float": "Critical",
                "Near-Critical":  "Near-Critical",
                "Float":          "Has Float",
            }).fillna("Has Float")

            fig = px.timeline(
                gantt_data.head(80),
                x_start="eff_start",
                x_end="eff_finish",
                y="Label",
                color="Bar Colour",
                color_discrete_map={
                    "Critical":     "#dc2626",
                    "Near-Critical":"#d97706",
                    "Has Float":    "#2563eb",
                },
                title=f"Trace Timeline: {trace_lbl}",
                labels={"Label": ""},
            )
            fig.update_yaxes(autorange="reversed")
            fig.update_layout(
                legend_title_text="Float Status",
                height=max(300, min(700, 60 + len(gantt_data.head(80)) * 28)),
                margin=dict(l=10, r=10, t=40, b=10),
            )

            # Add a vertical line for today
            fig.add_vline(
                x=datetime.now(),
                line_dash="dot",
                line_color="#6b7280",
                annotation_text="Today",
                annotation_position="top left",
            )

            st.plotly_chart(fig, use_container_width=True)
            if len(gantt_data) > 80:
                st.caption(f"Gantt shows first 80 of {len(gantt_data)} activities for performance.")
        else:
            st.info("No date data available to render a Gantt chart for this trace.")

        # ---- EXPORT ---------------------------------------------------------
        st.divider()

        # Build a summary sheet
        summary_data = {
            "Item":  ["Selected Activity", "Activity Name", "Trace Type",
                      "Total in chain", "Critical in chain",
                      "Near-Critical in chain", "Negative Float in chain"],
            "Value": [sel_code, sel_name, trace_lbl,
                      n_res, n_crit, n_near, n_neg],
        }

        export_sheets = {
            "Summary":    pd.DataFrame(summary_data),
            "Trace":      trace_df,
        }

        # Add filtered sub-sheets if there are enough rows
        crit_exp = trace_df[trace_df["Critical Flag"].isin(["Critical","Negative Float"])]
        near_exp = trace_df[trace_df["Critical Flag"] == "Near-Critical"]
        if not crit_exp.empty:
            export_sheets["Critical Activities"] = crit_exp
        if not near_exp.empty:
            export_sheets["Near-Critical"]       = near_exp

        xls_bytes = export_df_to_excel(export_sheets)

        st.download_button(
            label="📥  Export Logic Trace to Excel",
            data=xls_bytes,
            file_name=f"logic_trace_{sel_code}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            help="Exports the full trace table, critical activities and a summary sheet.",
        )

    else:
        # No trace run yet - show a prompt
        st.markdown(
            '<div style="background:#f0f9ff;border:1px dashed #93c5fd;border-radius:8px;'
            'padding:24px;text-align:center;color:#1e40af;margin-top:16px;">'
            '<strong>Press one of the buttons above to run a trace.</strong><br>'
            '<span style="font-size:13px;color:#6b7280;">Results will appear here.</span>'
            '</div>',
            unsafe_allow_html=True,
        )


# -----------------------------------------------------------------------------
# PAGE: CRITICAL PATH ANALYSIS
# -----------------------------------------------------------------------------

def page_critical_path(data: dict, near_crit_days: float):
    st.title("🚨 Critical Path Analysis")

    tasks = data["tasks_df"]
    rels = data["relationships_df"]

    if tasks.empty:
        st.warning("No activities loaded.")
        return

    tasks = get_critical_threshold(tasks, near_crit_days)

    tab1, tab2, tab3, tab4 = st.tabs(
        ["Critical Activities", "Near-Critical", "Negative Float", "By WBS / Package"]
    )

    with tab1:
        critical = tasks[tasks["is_critical"]].sort_values("total_float_days")
        st.metric("Critical Activities", len(critical))
        disp = ["task_code", "task_name", "wbs_path", "eff_start", "eff_finish",
                "total_float_days", "status"]
        avail = [c for c in disp if c in critical.columns]
        st.dataframe(critical[avail], use_container_width=True)

        if not critical.empty and "eff_start" in critical.columns:
            st.subheader("Critical Path Gantt")
            gantt_df = critical.dropna(subset=["eff_start", "eff_finish"]).copy()
            gantt_df["Start"] = gantt_df["eff_start"]
            gantt_df["Finish"] = gantt_df["eff_finish"]
            gantt_df["Task"] = gantt_df["task_code"] + " - " + gantt_df["task_name"]
            if len(gantt_df) > 0:
                fig = px.timeline(
                    gantt_df.head(50),
                    x_start="Start", x_end="Finish", y="Task",
                    title="Critical Path Activities (top 50)",
                    color_discrete_sequence=["#dc2626"],
                )
                fig.update_yaxes(autorange="reversed")
                st.plotly_chart(fig, use_container_width=True)

        xls = export_df_to_excel({"Critical Path": critical[avail]})
        st.download_button("📥 Export Critical Path", xls,
                           "critical_path.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    with tab2:
        near_crit = tasks[tasks["is_near_critical"]].sort_values("total_float_days")
        st.metric(f"Near-Critical (0 < float <= {near_crit_days}d)", len(near_crit))
        avail = [c for c in ["task_code","task_name","wbs_path","eff_start","eff_finish",
                              "total_float_days","status"] if c in near_crit.columns]
        st.dataframe(near_crit[avail], use_container_width=True)

    with tab3:
        neg = tasks[tasks["total_float_days"].apply(lambda f: f is not None and f < 0)].sort_values("total_float_days")
        st.metric("Negative Float Activities", len(neg))
        if not neg.empty:
            st.warning("⚠️ Activities with negative float indicate the schedule cannot be met -- investigate immediately.")
            avail = [c for c in ["task_code","task_name","total_float_days","eff_start",
                                  "eff_finish","status"] if c in neg.columns]
            st.dataframe(neg[avail], use_container_width=True)

    with tab4:
        if "wbs_path" not in tasks.columns:
            st.info("WBS data not available.")
            return
        tasks["wbs_top"] = tasks["wbs_path"].apply(
            lambda x: str(x).split(" > ")[0] if pd.notna(x) else "Unknown"
        )
        wbs_crit = tasks.groupby("wbs_top").agg(
            total=("task_id", "count"),
            critical=("is_critical", "sum"),
            near_critical=("is_near_critical", "sum"),
        ).reset_index()
        wbs_crit["crit_%"] = (wbs_crit["critical"] / wbs_crit["total"] * 100).round(1)
        fig = px.bar(
            wbs_crit, x="wbs_top", y=["critical", "near_critical"],
            title="Critical & Near-Critical by WBS",
            labels={"value": "Activities", "wbs_top": "WBS"},
            color_discrete_map={"critical": "#dc2626", "near_critical": "#f59e0b"},
            barmode="group",
        )
        st.plotly_chart(fig, use_container_width=True)
        st.dataframe(wbs_crit, use_container_width=True)


# -----------------------------------------------------------------------------
# PAGE: CRITICAL PATH TO SELECTED ACTIVITY
# -----------------------------------------------------------------------------

def _network_diagram_html(
    path_ids: list,
    all_pred_ids: list,
    task_lookup: dict,
    rels_df: pd.DataFrame,
) -> str:
    """
    Build a lightweight SVG-based network diagram showing the driving path
    as a horizontal chain, with branch predecessors shown above/below.
    Returns an HTML string ready for st.components.v1.html().
    """
    if not path_ids:
        return ""

    # Colour helpers
    def node_colour(tid):
        t  = task_lookup.get(tid, {})
        tf = safe_float(t.get("total_float_days"), 9999)
        if tf < 0:  return "#7f1d1d", "#fecaca"   # bg, border
        if tf == 0: return "#dc2626", "#fee2e2"
        if tf <= 10:return "#d97706", "#fef3c7"
        return       "#1d4ed8", "#dbeafe"

    BOX_W, BOX_H = 160, 54
    H_GAP, V_GAP = 40, 70
    n = len(path_ids)
    canvas_w = n * (BOX_W + H_GAP) + H_GAP
    canvas_h = 200

    svg_parts = [
        f'<svg xmlns="http://www.w3.org/2000/svg" '
        f'width="{canvas_w}" height="{canvas_h}" '
        f'style="font-family:Arial,sans-serif;font-size:11px;">'
    ]

    # Draw boxes for driving path (middle row y=70)
    mid_y = 80
    box_centres = {}
    for i, tid in enumerate(path_ids):
        t    = task_lookup.get(tid, {})
        x    = H_GAP + i * (BOX_W + H_GAP)
        y    = mid_y
        cx   = x + BOX_W // 2
        cy   = y + BOX_H // 2
        box_centres[tid] = (cx, cy)
        fc, bc = node_colour(tid)
        code = str(t.get("task_code", tid))[:14]
        name = str(t.get("task_name", ""))[:20]
        tf   = t.get("total_float_days")
        tf_s = f"Float: {tf}d" if tf is not None else ""
        svg_parts.append(
            f'<rect x="{x}" y="{y}" width="{BOX_W}" height="{BOX_H}" '
            f'rx="6" fill="{bc}" stroke="{fc}" stroke-width="2"/>'
            f'<text x="{cx}" y="{y+16}" text-anchor="middle" '
            f'font-weight="bold" fill="{fc}">{code}</text>'
            f'<text x="{cx}" y="{y+30}" text-anchor="middle" fill="#374151">{name}</text>'
            f'<text x="{cx}" y="{y+44}" text-anchor="middle" fill="{fc}" '
            f'font-size="10">{tf_s}</text>'
        )

        # Arrow from previous box
        if i > 0:
            prev_tid = path_ids[i - 1]
            px2, py2 = box_centres[prev_tid]
            # Get rel type
            rel_label = "FS"
            lag_label = ""
            if not rels_df.empty:
                rel = rels_df[
                    (rels_df.get("pred_task_id", pd.Series(dtype=str)) == prev_tid) &
                    (rels_df.get("succ_task_id", pd.Series(dtype=str)) == tid)
                ]
                if not rel.empty:
                    rel_label = str(rel["rel_type"].iloc[0])[-2:] if "rel_type" in rel.columns else "FS"
                    lag = safe_float(rel["lag_days"].iloc[0] if "lag_days" in rel.columns else 0, 0)
                    lag_label = f" +{int(lag)}d" if lag > 0 else (f" {int(lag)}d" if lag < 0 else "")
            ax1 = px2 + BOX_W // 2
            ax2 = x
            ay  = cy
            svg_parts.append(
                f'<line x1="{ax1}" y1="{ay}" x2="{ax2}" y2="{ay}" '
                f'stroke="#6b7280" stroke-width="2" marker-end="url(#arr)"/>'
                f'<text x="{(ax1+ax2)//2}" y="{ay-5}" text-anchor="middle" '
                f'fill="#6b7280" font-size="10">{rel_label}{lag_label}</text>'
            )

    # Arrow marker def
    svg_parts.insert(1,
        '<defs><marker id="arr" markerWidth="8" markerHeight="8" '
        'refX="6" refY="3" orient="auto">'
        '<path d="M0,0 L0,6 L8,3 z" fill="#6b7280"/>'
        '</marker></defs>'
    )

    svg_parts.append('</svg>')
    html = (
        '<div style="overflow-x:auto;background:#f8fafc;border:1px solid #e2e8f0;'
        'border-radius:8px;padding:12px;">'
        + "".join(svg_parts) +
        '</div>'
    )
    return html


def page_critical_path_to_activity(data: dict, near_crit_days: float):
    """
    Critical Path to Selected Activity page.

    Traces backwards through predecessor logic to identify the most likely
    driving chain into any selected activity or milestone.
    Uses float, finish dates, relationship type and lag to determine the driver.
    """
    st.title("🎯 Critical Path to Selected Activity")

    # -------------------------------------------------------------------------
    # EXPLANATION BANNER
    # -------------------------------------------------------------------------
    st.markdown(
        '<div style="background:#eff6ff;border-left:4px solid #3b82f6;'
        'border-radius:6px;padding:14px 18px;margin-bottom:18px;">'
        '<strong>How this works</strong><br>'
        'This shows the likely chain of activities driving the selected activity. '
        'It traces backwards through predecessor logic and identifies the most critical '
        'path based on total float, latest finish dates, relationship types and lag. '
        '<br><br>'
        '<em>This is based on available XER logic and float values and should be '
        'reviewed with the planner before being used for decision-making.</em>'
        '</div>',
        unsafe_allow_html=True,
    )

    tasks = data["tasks_df"]
    rels  = data["relationships_df"]

    # Guard rails
    if tasks.empty:
        st.warning("No activities found. Please upload an XER file first.")
        return
    if rels.empty:
        st.warning(
            "No relationship data found in this XER file. "
            "This page requires both activities and relationships."
        )
        return

    tasks = get_critical_threshold(tasks, near_crit_days)
    G           = build_graph(tasks, rels)
    task_lookup = tasks.set_index("task_id").to_dict("index")

    # -------------------------------------------------------------------------
    # ACTIVITY SELECTOR
    # -------------------------------------------------------------------------
    def _label(r):
        code = str(r.get("task_code", "?"))
        name = str(r.get("task_name", "?"))
        tf   = r.get("total_float_days")
        flag = " [CRITICAL]" if (tf is not None and safe_float(tf, 1) <= 0) else ""
        return f"{code}  --  {name}{flag}"

    act_labels = tasks.apply(_label, axis=1).tolist()
    selected_label = st.selectbox(
        "Select target activity or milestone",
        options=act_labels,
        key="cpta_selector",
        help="Choose the activity you want to understand the driving path for.",
    )
    sel_idx    = act_labels.index(selected_label)
    target_row = tasks.iloc[sel_idx]
    target_id  = target_row["task_id"]
    tgt_code   = str(target_row.get("task_code", "-"))
    tgt_name   = str(target_row.get("task_name", "-"))
    tgt_tf     = safe_float(target_row.get("total_float_days"), None) if "total_float_days" in target_row.index else None
    tgt_fcol   = _float_color(tgt_tf)
    tgt_crit   = bool(target_row.get("is_critical", False)) if "is_critical" in target_row.index else False
    tgt_stat   = _status_label(str(target_row.get("status", "")))
    tgt_scol   = _status_colour(str(target_row.get("status", "")))

    # Clear cached results when target changes
    if st.session_state.get("_cpta_last_id") != target_id:
        for k in ("cpta_path", "cpta_all_preds"):
            st.session_state.pop(k, None)
        st.session_state["_cpta_last_id"] = target_id

    # -------------------------------------------------------------------------
    # TARGET ACTIVITY BANNER
    # -------------------------------------------------------------------------
    crit_pill = (
        '<span style="background:#dc2626;color:white;padding:2px 10px;'
        'border-radius:12px;font-size:11px;font-weight:700;margin-left:8px;">CRITICAL</span>'
        if tgt_crit else ""
    )
    st.markdown(
        f"""
        <div style="background:#1e3a5f;color:white;border-radius:10px;
                    padding:16px 22px;margin:8px 0 18px 0;">
            <div style="font-size:12px;color:#93c5fd;font-weight:600;
                        letter-spacing:1px;text-transform:uppercase;">Target Activity</div>
            <div style="font-size:20px;font-weight:700;margin-top:4px;">
                {tgt_code}{crit_pill}
            </div>
            <div style="font-size:14px;color:#bfdbfe;margin-top:2px;">{tgt_name}</div>
            <div style="margin-top:10px;">
                <span style="background:{tgt_scol};color:white;padding:3px 10px;
                             border-radius:12px;font-size:12px;">{tgt_stat}</span>
                <span style="background:{tgt_fcol};color:white;padding:3px 10px;
                             border-radius:12px;font-size:12px;margin-left:6px;">
                    Float: {tgt_tf if tgt_tf is not None else "-"} days
                </span>
                <span style="color:#93c5fd;font-size:12px;margin-left:12px;">
                    Finish: {format_date(target_row.get("eff_finish") if "eff_finish" in target_row.index else None)}
                </span>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # Check if target has any predecessors at all
    direct_preds = list(G.predecessors(target_id))
    if not direct_preds:
        st.warning(
            f"**{tgt_code} has no predecessors.** This activity has an open start "
            "and is not driven by any logic in the programme. "
            "Nothing can be identified as the driving path."
        )
        return

    # -------------------------------------------------------------------------
    # RUN BUTTON
    # -------------------------------------------------------------------------
    run_col, _ = st.columns([1, 3])
    run_btn = run_col.button(
        "🔍  Find Driving Path",
        key="cpta_run",
        use_container_width=True,
        type="primary",
    )

    if run_btn:
        with st.spinner("Tracing predecessor network..."):
            driving_path   = driving_path_to_activity(G, tasks, rels, target_id)
            all_pred_pairs = trace_predecessors(G, target_id)
            all_pred_ids   = [p for p, _ in all_pred_pairs]
        st.session_state["cpta_path"]      = driving_path
        st.session_state["cpta_all_preds"] = all_pred_ids

    # -------------------------------------------------------------------------
    # RESULTS
    # -------------------------------------------------------------------------
    if "cpta_path" not in st.session_state:
        st.markdown(
            '<div style="background:#f0f9ff;border:1px dashed #93c5fd;border-radius:8px;'
            'padding:24px;text-align:center;color:#1e40af;margin-top:16px;">'
            '<strong>Press "Find Driving Path" above to run the analysis.</strong>'
            '</div>',
            unsafe_allow_html=True,
        )
        return

    driving_path = st.session_state["cpta_path"]
    all_pred_ids = st.session_state["cpta_all_preds"]

    # ---- KEY METRICS --------------------------------------------------------
    chain_tasks = tasks[tasks["task_id"].isin(driving_path)]
    n_chain     = len(driving_path)
    min_float   = chain_tasks["total_float_days"].min() if "total_float_days" in chain_tasks.columns else None
    n_crit_chain = int((chain_tasks["total_float_days"].apply(
        lambda f: safe_float(f, 1) <= 0
    )).sum()) if "total_float_days" in chain_tasks.columns else 0
    n_neg_chain  = int((chain_tasks["total_float_days"].apply(
        lambda f: safe_float(f, 0) < 0
    )).sum()) if "total_float_days" in chain_tasks.columns else 0

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Activities in Driving Chain",  n_chain)
    m2.metric("Lowest Float in Chain",         f"{min_float:.1f} days" if min_float is not None else "-")
    m3.metric("Critical in Chain",             n_crit_chain)
    m4.metric("Total Predecessor Network",     len(all_pred_ids))

    if n_neg_chain > 0:
        st.error(
            f"⚠️ **{n_neg_chain} activit{'y' if n_neg_chain == 1 else 'ies'} with negative float** "
            "in the driving chain. The current schedule cannot meet its target dates for this path."
        )

    st.divider()

    # ---- TABS ---------------------------------------------------------------
    tab_path, tab_network, tab_all_preds, tab_constraints = st.tabs([
        "Driving Path", "Network Diagram", "All Predecessors", "Constraints & Issues"
    ])

    # =========================================================================
    # TAB 1: DRIVING PATH TABLE
    # =========================================================================
    with tab_path:
        st.markdown(
            "The table below shows the most likely chain of activities driving "
            f"**{tgt_code}**, ordered from the earliest activity to the target. "
            "Activities are selected based on lowest float, latest finish date "
            "and relationship constraints."
        )

        path_rows = []
        for i, tid in enumerate(driving_path):
            t         = task_lookup.get(tid, {})
            tf        = t.get("total_float_days")
            is_target = (tid == target_id)

            # Relationship to next activity in chain
            rel_label = "-"
            lag_val   = 0
            if i < len(driving_path) - 1:
                next_tid = driving_path[i + 1]
                if not rels.empty:
                    rel = rels[
                        (rels.get("pred_task_id", pd.Series(dtype=str)) == tid) &
                        (rels.get("succ_task_id", pd.Series(dtype=str)) == next_tid)
                    ]
                    if not rel.empty:
                        rel_label = _rel_label(rel["rel_type"].iloc[0] if "rel_type" in rel.columns else "FS")
                        lag_val   = safe_float(rel["lag_days"].iloc[0] if "lag_days" in rel.columns else 0, 0)

            cstr = str(t.get("cstr_type", "")) if "cstr_type" in t else ""
            has_cstr = cstr.strip() not in ("", "None", "nan")

            path_rows.append({
                "Step":            i + 1,
                "Activity ID":     t.get("task_code", tid),
                "Activity Name":   t.get("task_name", ""),
                "Start":           format_date(t.get("eff_start")),
                "Finish":          format_date(t.get("eff_finish")),
                "Orig Dur (d)":    t.get("orig_dur_days", "-"),
                "Total Float (d)": tf if tf is not None else "-",
                "Link to Next":    rel_label if not is_target else "-",
                "Lag (d)":         lag_val if not is_target else "-",
                "Critical Flag":   _crit_flag(tf),
                "Constraint":      cstr if has_cstr else "",
                "Status":          _status_label(str(t.get("status", ""))),
                "Target":          "TARGET" if is_target else "",
            })

        path_df = pd.DataFrame(path_rows)

        # Colour code
        def _style_path_row(row):
            flag = row.get("Critical Flag", "")
            is_tgt = row.get("Target", "") == "TARGET"
            if is_tgt:
                return ["background-color:#1e3a5f;color:white;font-weight:700;"] * len(row)
            colour_map = {
                "Negative Float": "background-color:#fecaca;",
                "Critical":       "background-color:#fee2e2;",
                "Near-Critical":  "background-color:#fef3c7;",
            }
            style = colour_map.get(flag, "")
            return [style] * len(row)

        styled_path = path_df.style.apply(_style_path_row, axis=1)
        st.dataframe(styled_path, use_container_width=True, hide_index=True)

        # Gantt for driving path
        st.markdown("**Driving Path Timeline**")
        gantt_src = chain_tasks.dropna(subset=["eff_start","eff_finish"]).copy() if "eff_start" in chain_tasks.columns else pd.DataFrame()
        if not gantt_src.empty:
            gantt_src = gantt_src.merge(
                tasks[["task_id","task_code","task_name"]],
                on="task_id", how="left", suffixes=("","_t")
            )
            gantt_src["Label"]   = gantt_src["task_code"].astype(str) + "  " + gantt_src["task_name"].astype(str).str[:35]
            gantt_src["Colour"]  = gantt_src["task_id"].apply(
                lambda t: "Target" if t == target_id else (
                    "Critical" if safe_float(task_lookup.get(t,{}).get("total_float_days"), 1) <= 0
                    else "Near-Critical" if safe_float(task_lookup.get(t,{}).get("total_float_days"), 11) <= 10
                    else "Has Float"
                )
            )
            fig = px.timeline(
                gantt_src,
                x_start="eff_start", x_end="eff_finish", y="Label",
                color="Colour",
                color_discrete_map={
                    "Target":       "#1e3a5f",
                    "Critical":     "#dc2626",
                    "Near-Critical":"#d97706",
                    "Has Float":    "#2563eb",
                },
                title=f"Driving Path to {tgt_code}",
            )
            fig.update_yaxes(autorange="reversed")
            fig.add_vline(
                x=datetime.now(), line_dash="dot", line_color="#6b7280",
                annotation_text="Today", annotation_position="top left",
            )
            fig.update_layout(
                height=max(280, 50 + len(gantt_src) * 30),
                margin=dict(l=10, r=10, t=40, b=10),
                legend_title_text="Float Status",
            )
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("No date data available for the Gantt chart.")

    # =========================================================================
    # TAB 2: NETWORK DIAGRAM
    # =========================================================================
    with tab_network:
        st.markdown(
            "A simple left-to-right network diagram of the driving path. "
            "Each box shows the activity ID, name and float. "
            "Colours: **red** = critical/negative float, **amber** = near-critical, **blue** = has float, "
            "**navy** = target activity."
        )

        if len(driving_path) > 0:
            diagram_html = _network_diagram_html(
                driving_path, all_pred_ids, task_lookup, rels
            )
            if diagram_html:
                import streamlit.components.v1 as components
                n_boxes = len(driving_path)
                diagram_w = n_boxes * 200 + 80
                components.html(diagram_html, height=220, scrolling=True)
            else:
                st.info("Could not generate network diagram.")

            st.caption(
                "Note: The diagram shows the identified driving path only. "
                "Use the All Predecessors tab to see the full predecessor network."
            )
        else:
            st.info("No path data to display.")

    # =========================================================================
    # TAB 3: ALL PREDECESSORS
    # =========================================================================
    with tab_all_preds:
        all_pred_tasks = tasks[tasks["task_id"].isin(all_pred_ids)].copy()

        if all_pred_tasks.empty:
            st.info("No predecessor activities found.")
        else:
            n_ap     = len(all_pred_tasks)
            n_ap_crit = int((all_pred_tasks["total_float_days"].apply(
                lambda f: safe_float(f, 1) <= 0
            )).sum()) if "total_float_days" in all_pred_tasks.columns else 0

            st.markdown(
                _summary_bar(f"total predecessors", n_ap, "#374151") +
                (_summary_bar("critical", n_ap_crit, "#dc2626") if n_ap_crit else ""),
                unsafe_allow_html=True,
            )
            st.markdown("<br>", unsafe_allow_html=True)
            st.caption(
                "All activities in the predecessor network of the target activity, "
                "sorted by float (most critical first)."
            )

            all_pred_tasks = all_pred_tasks.sort_values("total_float_days")

            AP_COLS = {
                "task_code":        "Activity ID",
                "task_name":        "Activity Name",
                "wbs_path":         "WBS",
                "eff_start":        "Start",
                "eff_finish":       "Finish",
                "total_float_days": "Float (d)",
                "status":           "Status",
                "is_critical":      "Critical",
            }
            ap_show = {k: v for k, v in AP_COLS.items() if k in all_pred_tasks.columns}
            ap_df   = all_pred_tasks[list(ap_show.keys())].rename(columns=ap_show).copy()
            for col in ["Start","Finish"]:
                if col in ap_df.columns:
                    ap_df[col] = ap_df[col].apply(format_date)
            if "Critical" in ap_df.columns:
                ap_df["Critical"] = ap_df["Critical"].apply(lambda x: "Yes" if x else "")
            if "Status" in ap_df.columns:
                ap_df["Status"] = ap_df["Status"].apply(_status_label)

            st.dataframe(ap_df, use_container_width=True, hide_index=True, height=400)

    # =========================================================================
    # TAB 4: CONSTRAINTS & ISSUES
    # =========================================================================
    with tab_constraints:
        st.markdown(
            "Activities in the driving path or predecessor network that have "
            "constraints, negative float, or other schedule quality issues."
        )

        issues_found = False

        # --- Negative float in driving path ---
        neg_in_path = chain_tasks[
            chain_tasks["total_float_days"].apply(lambda f: safe_float(f, 0) < 0)
        ] if "total_float_days" in chain_tasks.columns else pd.DataFrame()

        if not neg_in_path.empty:
            issues_found = True
            st.markdown(
                '<div style="background:#fef2f2;border-left:4px solid #dc2626;'
                'border-radius:6px;padding:10px 14px;margin-bottom:12px;">'
                f'<strong>⚠️ Negative Float in Driving Chain ({len(neg_in_path)} activities)</strong><br>'
                'These activities are beyond their target dates. '
                'The driving chain cannot currently meet its schedule.'
                '</div>',
                unsafe_allow_html=True,
            )
            neg_cols = {k: v for k, v in {
                "task_code":"Activity ID","task_name":"Activity Name",
                "total_float_days":"Float (d)","eff_finish":"Finish","status":"Status"
            }.items() if k in neg_in_path.columns}
            neg_disp = neg_in_path[list(neg_cols.keys())].rename(columns=neg_cols).copy()
            if "Finish" in neg_disp.columns:
                neg_disp["Finish"] = neg_disp["Finish"].apply(format_date)
            st.dataframe(neg_disp, use_container_width=True, hide_index=True)

        # --- Constraints in driving path ---
        if "cstr_type" in chain_tasks.columns:
            constrained = chain_tasks[
                chain_tasks["cstr_type"].apply(
                    lambda x: bool(x) and str(x).strip() not in ("","None","nan")
                )
            ]
            if not constrained.empty:
                issues_found = True
                st.markdown(
                    '<div style="background:#fffbeb;border-left:4px solid #f59e0b;'
                    'border-radius:6px;padding:10px 14px;margin-bottom:12px;">'
                    f'<strong>Constraints in Driving Chain ({len(constrained)} activities)</strong><br>'
                    'Constraints override schedule logic and can cause artificial float or '
                    'negative float. Each one should be reviewed with the planner.'
                    '</div>',
                    unsafe_allow_html=True,
                )
                cstr_cols = {k: v for k, v in {
                    "task_code":"Activity ID","task_name":"Activity Name",
                    "cstr_type":"Constraint Type","cstr_date":"Constraint Date",
                    "total_float_days":"Float (d)"
                }.items() if k in constrained.columns}
                cstr_disp = constrained[list(cstr_cols.keys())].rename(columns=cstr_cols).copy()
                if "Constraint Date" in cstr_disp.columns:
                    cstr_disp["Constraint Date"] = cstr_disp["Constraint Date"].apply(format_date)
                st.dataframe(cstr_disp, use_container_width=True, hide_index=True)

        # --- Constraints in full predecessor network ---
        all_pred_tasks_full = tasks[tasks["task_id"].isin(all_pred_ids)].copy()
        if "cstr_type" in all_pred_tasks_full.columns:
            all_constrained = all_pred_tasks_full[
                all_pred_tasks_full["cstr_type"].apply(
                    lambda x: bool(x) and str(x).strip() not in ("","None","nan")
                )
            ]
            if not all_constrained.empty:
                issues_found = True
                with st.expander(f"Constraints in Full Predecessor Network ({len(all_constrained)})"):
                    cstr_cols2 = {k: v for k, v in {
                        "task_code":"Activity ID","task_name":"Activity Name",
                        "cstr_type":"Constraint Type","cstr_date":"Constraint Date",
                        "total_float_days":"Float (d)"
                    }.items() if k in all_constrained.columns}
                    cstr_disp2 = all_constrained[list(cstr_cols2.keys())].rename(columns=cstr_cols2).copy()
                    if "Constraint Date" in cstr_disp2.columns:
                        cstr_disp2["Constraint Date"] = cstr_disp2["Constraint Date"].apply(format_date)
                    st.dataframe(cstr_disp2, use_container_width=True, hide_index=True)

        # --- High lag in driving path relationships ---
        if not rels.empty and "lag_days" in rels.columns:
            path_set  = set(driving_path)
            path_rels = rels[
                rels.get("pred_task_id", pd.Series(dtype=str)).isin(path_set) &
                rels.get("succ_task_id", pd.Series(dtype=str)).isin(path_set)
            ]
            high_lag = path_rels[
                path_rels["lag_days"].apply(lambda l: abs(safe_float(l, 0)) > 5)
            ] if not path_rels.empty else pd.DataFrame()

            if not high_lag.empty:
                issues_found = True
                st.markdown(
                    '<div style="background:#eff6ff;border-left:4px solid #3b82f6;'
                    'border-radius:6px;padding:10px 14px;margin-bottom:12px;">'
                    f'<strong>Significant Lag in Driving Path ({len(high_lag)} relationships)</strong><br>'
                    'Lag of more than 5 days can hide logic issues and affect float calculations.'
                    '</div>',
                    unsafe_allow_html=True,
                )
                lag_cols = {k: v for k, v in {
                    "pred_task_code":"From","pred_task_name":"From Name",
                    "succ_task_code":"To","succ_task_name":"To Name",
                    "rel_type":"Link","lag_days":"Lag (d)"
                }.items() if k in high_lag.columns}
                if lag_cols:
                    st.dataframe(
                        high_lag[list(lag_cols.keys())].rename(columns=lag_cols),
                        use_container_width=True, hide_index=True,
                    )

        if not issues_found:
            st.success(
                "No constraints, negative float or significant lag found in the driving path. "
                "The chain appears logically sound."
            )

    # =========================================================================
    # EXCEL EXPORT
    # =========================================================================
    st.divider()

    # Rebuild path_df for export (already built above in tab_path)
    export_path_df = pd.DataFrame(path_rows) if path_rows else pd.DataFrame()

    summary_rows = {
        "Item":  [
            "Target Activity ID", "Target Activity Name", "Target Finish",
            "Target Float (days)", "Activities in Driving Chain",
            "Lowest Float in Chain", "Critical in Chain", "Negative Float in Chain",
            "Total Predecessor Network",
        ],
        "Value": [
            tgt_code, tgt_name,
            format_date(target_row.get("eff_finish") if "eff_finish" in target_row.index else None),
            tgt_tf,
            n_chain, min_float, n_crit_chain, n_neg_chain, len(all_pred_ids),
        ],
    }

    export_sheets = {
        "Summary":        pd.DataFrame(summary_rows),
        "Driving Path":   export_path_df,
        "All Predecessors": ap_df if not all_pred_tasks.empty else pd.DataFrame(columns=["No data"]),
    }

    if not neg_in_path.empty:
        export_sheets["Negative Float"] = neg_disp
    if "cstr_type" in chain_tasks.columns and not constrained.empty:
        export_sheets["Constraints"] = cstr_disp

    xls_bytes = export_df_to_excel(export_sheets)

    dl_col, _ = st.columns([1, 3])
    dl_col.download_button(
        label="📥  Export Driving Path Report to Excel",
        data=xls_bytes,
        file_name=f"driving_path_{tgt_code}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        help="Exports Summary, Driving Path, All Predecessors, Negative Float and Constraints sheets.",
        use_container_width=True,
    )

# -----------------------------------------------------------------------------
# PAGE: LABOUR HISTOGRAM
# -----------------------------------------------------------------------------

def page_labour_histogram(data: dict):
    st.title("👷 Labour Histogram")

    task_res = data["task_resources_df"]
    tasks = data["tasks_df"]
    resources = data["resources_df"]

    if task_res.empty:
        st.markdown("""
        <div class="warn-box">
        ⚠️ <strong>No resource loading found in this XER file.</strong><br>
        This usually means the programme was not resourced in P6, or resource data was not exported.
        <br><br>
        You can upload a separate resource CSV or Excel file below.
        </div>
        """, unsafe_allow_html=True)

        st.subheader("Upload Resource Loading File")
        res_file = st.file_uploader("Upload CSV or Excel (columns: task_code, rsrc_name, target_qty, target_start, target_finish)", type=["csv","xlsx"])
        if res_file:
            try:
                if res_file.name.endswith(".csv"):
                    task_res = pd.read_csv(res_file)
                else:
                    task_res = pd.read_excel(res_file)
                for col in ["target_start","target_finish"]:
                    if col in task_res.columns:
                        task_res[col] = pd.to_datetime(task_res[col], errors="coerce")
                st.success(f"Loaded {len(task_res)} resource rows.")
            except Exception as e:
                st.error(f"Could not read resource file: {e}")
                return
        else:
            return

    # Merge with task info and resource names
    if not tasks.empty and "task_id" in task_res.columns:
        task_res = task_res.merge(
            tasks[["task_id","task_code","task_name","wbs_path","is_critical" if "is_critical" in tasks.columns else "task_id"]].drop_duplicates(),
            on="task_id", how="left", suffixes=("","_task")
        )
    if not resources.empty and "rsrc_id" in task_res.columns:
        task_res = task_res.merge(resources[["rsrc_id","rsrc_name"]], on="rsrc_id", how="left", suffixes=("","_res"))
        if "rsrc_name_res" in task_res.columns:
            task_res["rsrc_name"] = task_res["rsrc_name_res"].fillna(task_res.get("rsrc_name",""))

    # Expand resource loading to weekly intervals
    def expand_to_weeks(df):
        rows = []
        for _, r in df.iterrows():
            s = pd.to_datetime(r.get("target_start") or r.get("target_start_date"))
            e = pd.to_datetime(r.get("target_finish") or r.get("target_end_date"))
            if pd.isna(s) or pd.isna(e) or s > e:
                continue
            qty = safe_float(r.get("target_qty", 0), 0)
            if qty == 0:
                continue
            weeks = max(1, math.ceil((e - s).days / 7))
            qty_per_week = qty / weeks
            current = s
            for _ in range(weeks):
                rows.append({
                    "week": current.to_period("W").start_time,
                    "month": current.to_period("M").start_time,
                    "qty": qty_per_week,
                    "rsrc_name": r.get("rsrc_name","Unknown"),
                    "task_code": r.get("task_code",""),
                    "task_name": r.get("task_name",""),
                    "wbs_path": r.get("wbs_path",""),
                })
                current += timedelta(weeks=1)
        return pd.DataFrame(rows)

    weekly = expand_to_weeks(task_res)

    if weekly.empty:
        st.warning("Could not generate histogram -- resource dates or quantities may be missing.")
        return

    # Filters
    st.sidebar.divider()
    st.sidebar.subheader("Labour Filters")
    all_resources = sorted(weekly["rsrc_name"].unique().tolist())
    sel_res = st.sidebar.multiselect("Resource / Trade", all_resources, default=all_resources)
    if sel_res:
        weekly = weekly[weekly["rsrc_name"].isin(sel_res)]

    # Metrics
    c1, c2, c3 = st.columns(3)
    c1.metric("Total Planned Hours", f"{weekly['qty'].sum():,.0f}")
    weekly_totals = weekly.groupby("week")["qty"].sum()
    c2.metric("Peak Week (hrs)", f"{weekly_totals.max():,.0f}" if not weekly_totals.empty else "-")
    c3.metric("Average Week (hrs)", f"{weekly_totals.mean():,.0f}" if not weekly_totals.empty else "-")

    tab1, tab2, tab3, tab4 = st.tabs(["By Week", "By Month", "By Resource", "By WBS"])

    with tab1:
        weekly_sum = weekly.groupby("week")["qty"].sum().reset_index()
        fig = px.bar(weekly_sum, x="week", y="qty",
                     title="Labour Loading by Week (Hours)",
                     labels={"week":"Week","qty":"Hours"},
                     color_discrete_sequence=["#2563eb"])
        st.plotly_chart(fig, use_container_width=True)

    with tab2:
        monthly_sum = weekly.groupby("month")["qty"].sum().reset_index()
        fig = px.bar(monthly_sum, x="month", y="qty",
                     title="Labour Loading by Month (Hours)",
                     labels={"month":"Month","qty":"Hours"},
                     color_discrete_sequence=["#1e3a5f"])
        st.plotly_chart(fig, use_container_width=True)

    with tab3:
        res_sum = weekly.groupby("rsrc_name")["qty"].sum().reset_index().sort_values("qty", ascending=False)
        fig = px.bar(res_sum, x="rsrc_name", y="qty",
                     title="Total Hours by Resource / Trade",
                     labels={"rsrc_name":"Resource","qty":"Hours"},
                     color_discrete_sequence=["#7c3aed"])
        st.plotly_chart(fig, use_container_width=True)

        # By week and resource stacked
        if len(sel_res) <= 10:
            by_res_week = weekly.groupby(["week","rsrc_name"])["qty"].sum().reset_index()
            fig2 = px.bar(by_res_week, x="week", y="qty", color="rsrc_name",
                          title="Weekly Labour by Resource",
                          labels={"week":"Week","qty":"Hours","rsrc_name":"Resource"})
            st.plotly_chart(fig2, use_container_width=True)

    with tab4:
        if "wbs_path" in weekly.columns:
            weekly["wbs_top"] = weekly["wbs_path"].apply(
                lambda x: str(x).split(" > ")[0] if pd.notna(x) and x else "Unknown"
            )
            wbs_sum = weekly.groupby("wbs_top")["qty"].sum().reset_index().sort_values("qty", ascending=False)
            fig = px.bar(wbs_sum, x="qty", y="wbs_top", orientation="h",
                         title="Total Hours by WBS",
                         color_discrete_sequence=["#059669"])
            st.plotly_chart(fig, use_container_width=True)

    # Export
    xls = export_df_to_excel({
        "Weekly Labour": weekly.groupby(["week","rsrc_name"])["qty"].sum().reset_index(),
        "Monthly Labour": weekly.groupby(["month","rsrc_name"])["qty"].sum().reset_index(),
        "By Resource": res_sum,
    })
    st.download_button("📥 Export Labour Data", xls, "labour_histogram.xlsx",
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# -----------------------------------------------------------------------------
# PAGE: SCHEDULE HEALTH CHECK
# -----------------------------------------------------------------------------

def page_health_check(data: dict, near_crit_days: float):
    st.title("🩺 Schedule Health Check")
    st.markdown("> Automated quality checks to identify common schedule issues.")

    tasks = data["tasks_df"]
    rels = data["relationships_df"]

    if tasks.empty:
        st.warning("No activities loaded.")
        return

    tasks = get_critical_threshold(tasks, near_crit_days)

    # Build predecessor/successor sets
    tasks_with_pred = set()
    tasks_with_succ = set()
    if not rels.empty:
        tasks_with_pred = set(rels["succ_task_id"].dropna()) if "succ_task_id" in rels.columns else set()
        tasks_with_succ = set(rels["pred_task_id"].dropna()) if "pred_task_id" in rels.columns else set()

    # Define checks
    checks = []

    # 1. No predecessors (excl. milestones at start)
    no_pred = tasks[~tasks["task_id"].isin(tasks_with_pred)]
    checks.append({
        "Check": "No Predecessors",
        "Count": len(no_pred),
        "Severity": "⚠️ Warning",
        "Why It Matters": "Activities with no predecessors are open-ended. They cannot be driven by logic and may cause float calculation issues.",
        "df": no_pred,
    })

    # 2. No successors
    no_succ = tasks[~tasks["task_id"].isin(tasks_with_succ)]
    checks.append({
        "Check": "No Successors",
        "Count": len(no_succ),
        "Severity": "⚠️ Warning",
        "Why It Matters": "Activities with no successors are open-ended and may have artificially high float.",
        "df": no_succ,
    })

    # 3. Negative float
    neg_float = tasks[tasks["total_float_days"].apply(lambda f: f is not None and f < 0)]
    checks.append({
        "Check": "Negative Float",
        "Count": len(neg_float),
        "Severity": "🔴 Critical",
        "Why It Matters": "Negative float means the current schedule cannot meet its target dates. Immediate attention required.",
        "df": neg_float,
    })

    # 4. High float (> 60 days)
    high_float = tasks[tasks["total_float_days"].apply(lambda f: f is not None and f > 60)]
    checks.append({
        "Check": "Very High Float (>60 days)",
        "Count": len(high_float),
        "Severity": "ℹ️ Info",
        "Why It Matters": "Activities with very high float may have missing logic or may not be properly constrained.",
        "df": high_float,
    })

    # 5. Excessive duration (> 60 working days)
    excess_dur = tasks[tasks["orig_dur_days"].apply(lambda d: d is not None and d > 60)]
    checks.append({
        "Check": "Excessive Duration (>60 days)",
        "Count": len(excess_dur),
        "Severity": "⚠️ Warning",
        "Why It Matters": "Very long activities are difficult to control and should usually be broken down into smaller work packages.",
        "df": excess_dur,
    })

    # 6. Constraints
    constrained = tasks[tasks["cstr_type"].apply(
        lambda x: bool(x) and str(x).strip() not in ("", "None")
    )] if "cstr_type" in tasks.columns else pd.DataFrame()
    checks.append({
        "Check": "Constrained Activities",
        "Count": len(constrained),
        "Severity": "⚠️ Warning",
        "Why It Matters": "Constraints override schedule logic and can create artificial float or negative float. Each constraint should be justified.",
        "df": constrained,
    })

    # 7. Excessive lag (> 10 days)
    if not rels.empty and "lag_days" in rels.columns:
        high_lag = rels[rels["lag_days"].apply(lambda l: l is not None and abs(safe_float(l,0)) > 10)]
        checks.append({
            "Check": "Excessive Lag (|lag| > 10 days)",
            "Count": len(high_lag),
            "Severity": "⚠️ Warning",
            "Why It Matters": "Excessive lag can hide critical path issues. Lag should be replaced with properly sequenced activities.",
            "df": high_lag,
        })

    # 8. Missing dates
    missing_dates = tasks[tasks["eff_start"].isna() | tasks["eff_finish"].isna()]
    checks.append({
        "Check": "Missing Start or Finish Dates",
        "Count": len(missing_dates),
        "Severity": "🔴 Critical",
        "Why It Matters": "Activities with no dates cannot be scheduled or reported on.",
        "df": missing_dates,
    })

    # 9. Actual dates in future
    now = datetime.now()
    future_actuals = tasks[
        tasks["act_start"].apply(lambda d: d is not None and d > now) |
        tasks["act_finish"].apply(lambda d: d is not None and d > now)
    ] if "act_start" in tasks.columns else pd.DataFrame()
    checks.append({
        "Check": "Future Actual Dates",
        "Count": len(future_actuals),
        "Severity": "🔴 Critical",
        "Why It Matters": "Actual start/finish dates should not be in the future. This indicates data entry errors.",
        "df": future_actuals,
    })

    # 10. Critical not started
    crit_not_started = tasks[
        tasks["is_critical"] &
        tasks["status"].apply(lambda s: str(s) in ("TK_NotStart", "Not Started") if pd.notna(s) else False)
    ] if "status" in tasks.columns else pd.DataFrame()
    checks.append({
        "Check": "Critical Activities Not Started",
        "Count": len(crit_not_started),
        "Severity": "🔴 Critical",
        "Why It Matters": "Critical activities that haven't started need immediate attention to avoid slippage.",
        "df": crit_not_started,
    })

    # 11. Near-critical due in 8 weeks
    eight_weeks = now + timedelta(weeks=8)
    near_due = tasks[
        tasks["is_near_critical"] &
        tasks["eff_finish"].apply(lambda d: d is not None and d <= eight_weeks)
    ] if "eff_finish" in tasks.columns else pd.DataFrame()
    checks.append({
        "Check": "Near-Critical Due in 8 Weeks",
        "Count": len(near_due),
        "Severity": "⚠️ Warning",
        "Why It Matters": "Near-critical activities finishing soon may become critical if not progressed.",
        "df": near_due,
    })

    # Scorecard
    st.subheader("Health Check Scorecard")
    score_data = [
        {"Check": c["Check"], "Count": c["Count"], "Severity": c["Severity"]}
        for c in checks
    ]
    score_df = pd.DataFrame(score_data)
    st.dataframe(score_df, use_container_width=True)

    # Detail per check
    st.divider()
    for chk in checks:
        with st.expander(f"{chk['Severity']} -- {chk['Check']} ({chk['Count']})"):
            st.markdown(f"**Why it matters:** {chk['Why It Matters']}")
            df = chk["df"]
            if not df.empty:
                disp = [c for c in ["task_code","task_name","wbs_path","eff_start",
                                     "eff_finish","total_float_days","status",
                                     "cstr_type","lag_days"] if c in df.columns]
                st.dataframe(df[disp].head(100), use_container_width=True)
                # Export individual check
                xls = export_df_to_excel({chk["Check"][:31]: df[disp]})
                st.download_button(
                    f"📥 Export: {chk['Check']}", xls,
                    f"health_{chk['Check'][:20].replace(' ','_')}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            else:
                st.success("✅ No issues found for this check.")

    # Full export
    all_export = {chk["Check"][:31]: chk["df"][[c for c in ["task_code","task_name","total_float_days","status"] if c in chk["df"].columns]] if not chk["df"].empty else pd.DataFrame(columns=["No issues"]) for chk in checks}
    xls_all = export_df_to_excel(all_export)
    st.download_button("📥 Export Full Health Check Report", xls_all, "schedule_health_check.xlsx",
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# -----------------------------------------------------------------------------
# PAGE: PLANNING NOTES
# -----------------------------------------------------------------------------

HIGHLIGHT_WORDS = [
    "risk", "delay", "delayed", "blocked", "constraint", "access",
    "design", "procurement", "client", "instruction", "CE", "EWN",
    "change", "issue", "hold", "pending", "late", "overrun",
]

def highlight_text(text: str) -> str:
    """Wrap highlight words in HTML span."""
    for word in HIGHLIGHT_WORDS:
        pattern = re.compile(r"\b(" + re.escape(word) + r")\b", re.IGNORECASE)
        text = pattern.sub(r'<span style="background:#fef08a;font-weight:bold;">\1</span>', text)
    return text


def page_planning_notes(data: dict):
    st.title("📝 Planning Notes")
    st.markdown("> Upload planning notes and link them to activities in the programme.")

    tasks = data["tasks_df"]
    notes_file = st.file_uploader("Upload Planning Notes (CSV, Excel, TXT, or DOCX)",
                                   type=["csv","xlsx","txt","docx"])

    if notes_file is None:
        st.info("Upload a notes file to get started. The file should contain free-text notes referencing activity IDs.")
        return

    # Read notes
    notes_text = ""
    notes_rows = []

    try:
        if notes_file.name.endswith(".csv"):
            df = pd.read_csv(notes_file)
            notes_text = " ".join(df.astype(str).values.flatten())
            notes_rows = df.to_dict("records")
        elif notes_file.name.endswith(".xlsx"):
            df = pd.read_excel(notes_file)
            notes_text = " ".join(df.astype(str).values.flatten())
            notes_rows = df.to_dict("records")
        elif notes_file.name.endswith(".txt"):
            notes_text = notes_file.read().decode("utf-8", errors="replace")
            notes_rows = [{"line": i+1, "text": line} for i, line in enumerate(notes_text.splitlines()) if line.strip()]
        elif notes_file.name.endswith(".docx"):
            from docx import Document
            doc = Document(io.BytesIO(notes_file.read()))
            lines = [p.text for p in doc.paragraphs if p.text.strip()]
            notes_text = "\n".join(lines)
            notes_rows = [{"paragraph": i+1, "text": line} for i, line in enumerate(lines)]
        else:
            st.error("Unsupported file format.")
            return
        st.success(f"Loaded notes file: {notes_file.name}")
    except Exception as e:
        st.error(f"Could not read notes file: {e}")
        return

    # Find activity IDs mentioned in notes
    if not tasks.empty and "task_code" in tasks.columns:
        task_codes = tasks["task_code"].dropna().tolist()
        found_codes = [code for code in task_codes if code in notes_text]

        st.subheader(f"Activity IDs Found in Notes: {len(found_codes)}")
        if found_codes:
            matched_tasks = tasks[tasks["task_code"].isin(found_codes)][
                ["task_code","task_name","eff_start","eff_finish","total_float_days","status"]
            ]
            st.dataframe(matched_tasks, use_container_width=True)
        else:
            st.info("No activity IDs from the programme were found in the notes.")

        # Not found
        not_found = [code for code in task_codes if code not in notes_text]
        st.caption(f"{len(not_found)} activities not mentioned in notes.")

    # Keyword search
    st.divider()
    st.subheader("Keyword Search")
    keyword = st.text_input("Search notes for keyword")

    display_rows = notes_rows
    if keyword:
        display_rows = [r for r in notes_rows if keyword.lower() in str(r).lower()]
        st.caption(f"{len(display_rows)} matching entries")

    # Display with highlights
    for row in display_rows[:100]:
        text = str(row.get("text","") or list(row.values())[-1])
        highlighted = highlight_text(text)
        st.markdown(f"<div style='background:#f8fafc;border-left:3px solid #2563eb;padding:8px;margin:4px 0;font-size:13px;'>{highlighted}</div>", unsafe_allow_html=True)

    # Full highlighted dump
    st.divider()
    st.subheader("Full Notes (with keyword highlighting)")
    highlighted_full = highlight_text(notes_text.replace("\n","<br>"))
    st.markdown(f"<div style='background:white;border:1px solid #e2e8f0;padding:16px;border-radius:8px;max-height:400px;overflow-y:auto;font-size:12px;'>{highlighted_full}</div>", unsafe_allow_html=True)


# -----------------------------------------------------------------------------
# PAGE: PROGRAMME COMPARISON
# -----------------------------------------------------------------------------

def page_programme_comparison():
    st.title("📅 Programme Comparison")
    st.markdown("> Compare two programme revisions to identify changes in dates, float, and status.")

    col1, col2 = st.columns(2)
    with col1:
        st.subheader("Previous Programme")
        prev_file = st.file_uploader("Upload Previous XER", type=["xer"], key="prev_xer")
    with col2:
        st.subheader("Current Programme")
        curr_file = st.file_uploader("Upload Current XER", type=["xer"], key="curr_xer")

    if not prev_file or not curr_file:
        st.info("Upload both XER files above to compare programmes.")
        return

    with st.spinner("Parsing both programmes..."):
        prev_data = parse_xer(prev_file.read())
        curr_data = parse_xer(curr_file.read())

    prev_tasks = prev_data["tasks_df"]
    curr_tasks = curr_data["tasks_df"]

    if prev_tasks.empty or curr_tasks.empty:
        st.error("Could not parse one or both files.")
        return

    prev_tasks = get_critical_threshold(prev_tasks)
    curr_tasks = get_critical_threshold(curr_tasks)

    # Merge on task_code
    merged = prev_tasks.merge(
        curr_tasks, on="task_code", how="outer", suffixes=("_prev","_curr")
    )

    # Added / deleted
    added = curr_tasks[~curr_tasks["task_code"].isin(prev_tasks["task_code"])]
    deleted = prev_tasks[~prev_tasks["task_code"].isin(curr_tasks["task_code"])]

    # Changed activities
    common = merged.dropna(subset=["task_code"])

    def date_diff_days(d1, d2):
        if pd.isna(d1) or pd.isna(d2):
            return None
        try:
            return int((pd.Timestamp(d2) - pd.Timestamp(d1)).days)
        except Exception:
            return None

    common = common.copy()
    common["start_movement"] = common.apply(
        lambda r: date_diff_days(r.get("eff_start_prev"), r.get("eff_start_curr")), axis=1
    )
    common["finish_movement"] = common.apply(
        lambda r: date_diff_days(r.get("eff_finish_prev"), r.get("eff_finish_curr")), axis=1
    )
    common["float_movement"] = common.apply(
        lambda r: safe_float(r.get("total_float_days_curr"), 0) - safe_float(r.get("total_float_days_prev"), 0), axis=1
    )

    # Became critical / stopped being critical
    if "is_critical_prev" in common.columns and "is_critical_curr" in common.columns:
        became_crit = common[~common["is_critical_prev"].fillna(False) & common["is_critical_curr"].fillna(False)]
        stopped_crit = common[common["is_critical_prev"].fillna(False) & ~common["is_critical_curr"].fillna(False)]
    else:
        became_crit = pd.DataFrame()
        stopped_crit = pd.DataFrame()

    # Tabs
    tab1, tab2, tab3, tab4 = st.tabs(["Summary", "Added/Deleted", "Date Movement", "Critical Changes"])

    with tab1:
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Added Activities", len(added))
        c2.metric("Deleted Activities", len(deleted))
        c3.metric("Became Critical", len(became_crit))
        c4.metric("Stopped Being Critical", len(stopped_crit))

        slipped = common[common["finish_movement"].apply(lambda x: x is not None and x > 0)]
        brought_fwd = common[common["finish_movement"].apply(lambda x: x is not None and x < 0)]
        c1.metric("Finish Slipped", len(slipped))
        c2.metric("Finish Brought Forward", len(brought_fwd))

    with tab2:
        st.subheader(f"Added Activities ({len(added)})")
        if not added.empty:
            avail = [c for c in ["task_code","task_name","eff_start","eff_finish","total_float_days"] if c in added.columns]
            st.dataframe(added[avail], use_container_width=True)
        st.subheader(f"Deleted Activities ({len(deleted)})")
        if not deleted.empty:
            avail = [c for c in ["task_code","task_name","eff_start","eff_finish","total_float_days"] if c in deleted.columns]
            st.dataframe(deleted[avail], use_container_width=True)

    with tab3:
        st.subheader("Date & Float Movement")
        move_cols = ["task_code","task_name_curr","start_movement","finish_movement","float_movement",
                     "eff_start_prev","eff_start_curr","eff_finish_prev","eff_finish_curr"]
        avail = [c for c in move_cols if c in common.columns]
        st.dataframe(common[avail].sort_values("finish_movement", ascending=False, na_position="last"), use_container_width=True)

        if "finish_movement" in common.columns:
            fig = px.histogram(common["finish_movement"].dropna(), nbins=30,
                               title="Finish Date Movement Distribution (days, positive = slipped)",
                               color_discrete_sequence=["#2563eb"])
            fig.add_vline(x=0, line_dash="dash", line_color="green")
            st.plotly_chart(fig, use_container_width=True)

    with tab4:
        st.subheader(f"Became Critical ({len(became_crit)})")
        if not became_crit.empty:
            st.dataframe(became_crit[[c for c in ["task_code","task_name_curr","eff_finish_prev","eff_finish_curr","float_movement"] if c in became_crit.columns]], use_container_width=True)
        st.subheader(f"Stopped Being Critical ({len(stopped_crit)})")
        if not stopped_crit.empty:
            st.dataframe(stopped_crit[[c for c in ["task_code","task_name_curr","eff_finish_prev","eff_finish_curr","float_movement"] if c in stopped_crit.columns]], use_container_width=True)

    # Export
    xls = export_df_to_excel({
        "Added": added[[c for c in ["task_code","task_name","eff_start","eff_finish"] if c in added.columns]] if not added.empty else pd.DataFrame(columns=["No data"]),
        "Deleted": deleted[[c for c in ["task_code","task_name","eff_start","eff_finish"] if c in deleted.columns]] if not deleted.empty else pd.DataFrame(columns=["No data"]),
        "Date Movement": common[[c for c in move_cols if c in common.columns]],
        "Became Critical": became_crit[[c for c in ["task_code","task_name_curr","finish_movement"] if c in became_crit.columns]] if not became_crit.empty else pd.DataFrame(columns=["No data"]),
    })
    st.download_button("📥 Export Comparison Report", xls, "programme_comparison.xlsx",
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# -----------------------------------------------------------------------------
# PAGE: EXPORT REPORTS
# -----------------------------------------------------------------------------

def page_export_reports(data: dict, near_crit_days: float):
    st.title("📥 Export Reports")
    st.markdown("> Download all schedule data as formatted Excel reports.")

    tasks = data["tasks_df"]
    rels = data["relationships_df"]
    wbs = data["wbs_df"]
    resources = data["resources_df"]

    if tasks.empty:
        st.warning("No data loaded to export.")
        return

    tasks = get_critical_threshold(tasks, near_crit_days)
    critical = tasks[tasks["is_critical"]]
    neg_float = tasks[tasks["total_float_days"].apply(lambda f: f is not None and f < 0)]

    col1, col2 = st.columns(2)

    with col1:
        st.subheader("Single-Sheet Exports")

        # All activities
        avail = [c for c in ["task_code","task_name","wbs_path","eff_start","eff_finish",
                              "orig_dur_days","rem_dur_days","total_float_days","free_float_days",
                              "status","task_type","is_critical","cstr_type"] if c in tasks.columns]
        xls = export_df_to_excel({"All Activities": tasks[avail]})
        st.download_button("📄 All Activities", xls, "all_activities.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        # Critical path
        avail_c = [c for c in avail if c in critical.columns]
        xls2 = export_df_to_excel({"Critical Path": critical[avail_c]})
        st.download_button("🔴 Critical Path Activities", xls2, "critical_path.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        # Relationships
        if not rels.empty:
            xls3 = export_df_to_excel({"Relationships": rels})
            st.download_button("🔗 All Relationships", xls3, "relationships.xlsx",
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    with col2:
        st.subheader("Multi-Sheet Reports")

        # Full schedule pack
        sheets = {"All Activities": tasks[avail]}
        if not critical.empty:
            sheets["Critical Path"] = critical[avail_c]
        if not neg_float.empty:
            sheets["Negative Float"] = neg_float[[c for c in avail if c in neg_float.columns]]
        if not rels.empty:
            sheets["Relationships"] = rels
        if not wbs.empty:
            sheets["WBS"] = wbs
        if not resources.empty:
            sheets["Resources"] = resources

        xls_full = export_df_to_excel(sheets)
        st.download_button("📦 Full Schedule Data Pack", xls_full, "schedule_data_pack.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        # WBS summary
        if "wbs_path" in tasks.columns:
            tasks["wbs_top"] = tasks["wbs_path"].apply(
                lambda x: str(x).split(" > ")[0] if pd.notna(x) and x else "Unknown"
            )
            wbs_summary = tasks.groupby("wbs_top").agg(
                total=("task_id","count"),
                critical=("is_critical","sum"),
                near_critical=("is_near_critical","sum"),
            ).reset_index()
            xls_wbs = export_df_to_excel({"WBS Summary": wbs_summary})
            st.download_button("🌲 WBS Summary", xls_wbs, "wbs_summary.xlsx",
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# -----------------------------------------------------------------------------
# PAGE: HOME  (PlanTrace branded landing page)
# -----------------------------------------------------------------------------

def _page_home():
    """
    PlanTrace branded homepage.
    Shown when no XER is loaded, or when the user navigates to Home.
    """

    # ---- Hero section -------------------------------------------------------
    st.markdown(
        """
        <div style="padding: 48px 0 24px 0;">
            <div style="font-size: 13px; font-weight: 700; color: #F5A623;
                        letter-spacing: 2px; text-transform: uppercase;
                        margin-bottom: 10px;">
                Project Programme Intelligence
            </div>
            <div style="font-size: 52px; font-weight: 900; color: #0B1F33;
                        line-height: 1.1; letter-spacing: -1px;">
                PlanTrace
            </div>
            <div style="width: 56px; height: 4px; background: #F5A623;
                        border-radius: 2px; margin: 14px 0 18px 0;"></div>
            <div style="font-size: 20px; font-weight: 400; color: #334155;
                        margin-bottom: 10px;">
                Trace logic. Expose risk. Drive delivery.
            </div>
            <div style="font-size: 15px; color: #64748B; max-width: 680px;
                        line-height: 1.7; margin-bottom: 32px;">
                Project planning intelligence for delivery teams. Upload an XER programme,
                trace predecessors and successors, review critical paths, check programme
                health and understand labour demand &mdash; without opening P6.
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # ---- Upload prompt -------------------------------------------------------
    st.markdown(
        """
        <div style="background:#0B1F33; border-radius:12px; padding:22px 28px;
                    display:flex; align-items:center; gap:20px; margin-bottom:36px;
                    max-width:560px;">
            <div style="font-size:28px;">📂</div>
            <div>
                <div style="color:#F5A623;font-weight:700;font-size:15px;
                            margin-bottom:4px;">Ready to start</div>
                <div style="color:#CBD5E1;font-size:13px;line-height:1.5;">
                    Upload your <strong style="color:#fff;">.xer file</strong>
                    using the panel on the left to begin analysis.
                    <br>Export from P6 via
                    <strong style="color:#F5A623;">File &rarr; Export &rarr; Primavera P6 XER</strong>
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # ---- Feature cards -------------------------------------------------------
    st.markdown(
        '<div style="font-size:13px;font-weight:700;color:#94A3B8;'
        'letter-spacing:1.5px;text-transform:uppercase;margin-bottom:16px;">'
        'What PlanTrace does</div>',
        unsafe_allow_html=True,
    )

    CARDS = [
        {
            "icon": "🔗",
            "title": "Logic Trace",
            "body": (
                "See what drives an activity and what it impacts. "
                "Trace full predecessor and successor chains across the network, "
                "with depth levels and relationship types shown at every step."
            ),
        },
        {
            "icon": "🚨",
            "title": "Critical Path",
            "body": (
                "Review the full critical path, near-critical work and negative float. "
                "Identify which activity or milestone is at risk and understand "
                "exactly what is driving it."
            ),
        },
        {
            "icon": "👷",
            "title": "Labour Demand",
            "body": (
                "View labour histograms by week, month, WBS and resource. "
                "Identify peak demand periods and understand resource loading "
                "across the programme."
            ),
        },
        {
            "icon": "🩺",
            "title": "Programme Health",
            "body": (
                "Find missing logic, open ends, constraints, excessive lag and "
                "planning risk before they cause problems. "
                "Eleven automated quality checks with export."
            ),
        },
    ]

    cols = st.columns(4, gap="medium")
    for col, card in zip(cols, CARDS):
        with col:
            st.markdown(
                f"""
                <div class="pt-card">
                    <div class="pt-card-icon">{card["icon"]}</div>
                    <div class="pt-card-accent"></div>
                    <div class="pt-card-title">{card["title"]}</div>
                    <div class="pt-card-body">{card["body"]}</div>
                </div>
                """,
                unsafe_allow_html=True,
            )

    # ---- What's in the tool -------------------------------------------------
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown(
        '<div style="font-size:13px;font-weight:700;color:#94A3B8;'
        'letter-spacing:1.5px;text-transform:uppercase;margin-bottom:16px;">'
        'All pages</div>',
        unsafe_allow_html=True,
    )

    PAGE_LIST = [
        ("📊", "Project Summary",          "Activity counts, float distribution, WBS breakdown and schedule span."),
        ("🔍", "Activity Search",           "Search and filter activities. View full detail, dates, float and logic."),
        ("🔗", "Logic Trace",               "Trace predecessors and successors through the network with depth levels."),
        ("🚨", "Critical Path Analysis",    "Full critical path, near-critical and negative float by WBS."),
        ("🎯", "Critical Path to Activity", "Identify the driving chain into any selected activity or milestone."),
        ("👷", "Labour Histogram",          "Weekly and monthly labour demand by resource, WBS and package."),
        ("🩺", "Schedule Health Check",     "Eleven automated quality checks with counts, tables and export."),
        ("📝", "Planning Notes",            "Upload notes, link to activities, keyword search and highlighting."),
        ("📅", "Programme Comparison",      "Compare two XER revisions. See what moved, changed or became critical."),
        ("📥", "Export Reports",            "Download all data as formatted Excel workbooks."),
    ]

    left, right = st.columns(2, gap="large")
    for i, (icon, title, desc) in enumerate(PAGE_LIST):
        col = left if i % 2 == 0 else right
        with col:
            st.markdown(
                f"""
                <div style="display:flex;gap:14px;align-items:flex-start;
                            padding:14px 0;border-bottom:1px solid #E2E8F0;">
                    <div style="font-size:22px;min-width:30px;margin-top:2px;">{icon}</div>
                    <div>
                        <div style="font-weight:700;color:#0B1F33;
                                    font-size:14px;margin-bottom:3px;">{title}</div>
                        <div style="color:#64748B;font-size:13px;
                                    line-height:1.5;">{desc}</div>
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )

    # ---- Footer -------------------------------------------------------------
    st.markdown("<br><br>", unsafe_allow_html=True)
    st.markdown(
        """
        <div style="border-top:1px solid #E2E8F0;padding-top:18px;
                    display:flex;justify-content:space-between;align-items:center;">
            <div style="font-size:13px;color:#94A3B8;">
                <strong style="color:#0B1F33;">PlanTrace</strong>
                &nbsp;&nbsp;|&nbsp;&nbsp;
                Built for Primavera P6 XER programmes
                &nbsp;&nbsp;|&nbsp;&nbsp;
                No P6 licence required
            </div>
            <div style="font-size:12px;color:#CBD5E1;">
                Upload a .xer file to begin
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


# -----------------------------------------------------------------------------
# SIDEBAR & MAIN APP
# -----------------------------------------------------------------------------

def sidebar_upload():
    """Render the PlanTrace sidebar and return (xer_file, near_crit_days, page)."""
    with st.sidebar:
        # Brand header
        st.markdown(
            """
            <div style="padding:18px 4px 6px 4px;">
                <div style="font-size:22px;font-weight:800;color:#F5A623;
                            letter-spacing:1px;">PlanTrace</div>
                <div style="font-size:11px;color:#64748B;margin-top:2px;
                            letter-spacing:0.5px;text-transform:uppercase;">
                    Programme Intelligence
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        st.markdown(
            '<hr style="border:none;border-top:1px solid #1e3a5f;margin:4px 0 14px 0;">',
            unsafe_allow_html=True,
        )

        xer_file = st.file_uploader(
            "Upload XER File",
            type=["xer"],
            help="Export from P6: File > Export > Primavera P6 XER",
        )

        st.markdown(
            '<hr style="border:none;border-top:1px solid #1e3a5f;margin:14px 0 10px 0;">',
            unsafe_allow_html=True,
        )

        st.markdown(
            '<div style="font-size:11px;color:#94A3B8;font-weight:600;'
            'letter-spacing:0.8px;text-transform:uppercase;'
            'margin-bottom:6px;">Settings</div>',
            unsafe_allow_html=True,
        )
        near_crit_days = st.slider(
            "Near-Critical Float (days)",
            min_value=1, max_value=30, value=10, step=1,
            help="Activities with float between 0 and this value are flagged as near-critical.",
        )

        st.markdown(
            '<hr style="border:none;border-top:1px solid #1e3a5f;margin:14px 0 10px 0;">',
            unsafe_allow_html=True,
        )

        st.markdown(
            '<div style="font-size:11px;color:#94A3B8;font-weight:600;'
            'letter-spacing:0.8px;text-transform:uppercase;'
            'margin-bottom:6px;">Navigation</div>',
            unsafe_allow_html=True,
        )
        page = st.selectbox(
            "Go to page",
            [
                "🏠 Home",
                "📊 Project Summary",
                "🔍 Activity Search",
                "🔗 Logic Trace",
                "🚨 Critical Path Analysis",
                "🎯 Critical Path to Activity",
                "👷 Labour Histogram",
                "🩺 Schedule Health Check",
                "📝 Planning Notes",
                "📅 Programme Comparison",
                "📥 Export Reports",
            ],
            label_visibility="collapsed",
        )

        st.markdown(
            '<hr style="border:none;border-top:1px solid #1e3a5f;margin:14px 0 10px 0;">',
            unsafe_allow_html=True,
        )
        st.markdown(
            '<div style="font-size:11px;color:#475569;line-height:1.6;">'
            'Export from Primavera P6:<br>'
            '<strong style="color:#94A3B8;">File &rarr; Export &rarr; P6 XER</strong>'
            '</div>',
            unsafe_allow_html=True,
        )

    return xer_file, near_crit_days, page


def main():
    xer_file, near_crit_days, page = sidebar_upload()

    # Programme comparison doesn't need the main file loaded
    if page == "📅 Programme Comparison":
        page_programme_comparison()
        return

    # Load XER
    if xer_file is None or page == "🏠 Home":
        _page_home()
        if xer_file is None:
            return

    # Cache parsed data in session state
    cache_key = f"xer_data_{xer_file.name}_{xer_file.size}"
    if cache_key not in st.session_state:
        with st.spinner(f"Parsing {xer_file.name}..."):
            try:
                data = parse_xer(xer_file.read())
                st.session_state[cache_key] = data
                st.session_state["current_xer_key"] = cache_key
            except Exception as e:
                st.error(f"Failed to parse XER file: {e}")
                return
    else:
        data = st.session_state[cache_key]
        st.session_state["current_xer_key"] = cache_key

    # Show parse method info
    method = data.get("parse_method","-")
    n_tasks = len(data["tasks_df"])
    n_rels = len(data["relationships_df"])
    st.sidebar.success(f"✅ Loaded: {n_tasks} activities, {n_rels} relationships")
    st.sidebar.caption(f"Parser: {method}")

    # Route to pages
    if page == "🏠 Home":
        _page_home()
        return
    elif page == "📊 Project Summary":
        page_project_summary(data, near_crit_days)
    elif page == "🔍 Activity Search":
        page_activity_search(data, near_crit_days)
    elif page == "🔗 Logic Trace":
        page_logic_trace(data, near_crit_days)
    elif page == "🚨 Critical Path Analysis":
        page_critical_path(data, near_crit_days)
    elif page == "🎯 Critical Path to Activity":
        page_critical_path_to_activity(data, near_crit_days)
    elif page == "👷 Labour Histogram":
        page_labour_histogram(data)
    elif page == "🩺 Schedule Health Check":
        page_health_check(data, near_crit_days)
    elif page == "📝 Planning Notes":
        page_planning_notes(data)
    elif page == "📥 Export Reports":
        page_export_reports(data, near_crit_days)


if __name__ == "__main__":
    main()
